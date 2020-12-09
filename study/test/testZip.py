import configparser
import os
import sys
import time
import xml.etree.ElementTree as ET
import zipfile
import pandas as pd
from openpyxl import load_workbook
from pandas import DataFrame

def deal_database_param(databases=None, params=None, systemType=None, appType=None, appName=None, nodeId=None):
    if databases is None or params is None:
        return
    if databases is not None:
        database_list = databases.findall("database")
        if database_list is not None and len(database_list) > 0:
            for database in database_list:
                if database.attrib.get("id") is not None:
                    auth = database.find("auth")
                    if auth is not None:
                        user = auth.attrib.get("user")
                        if user is None:
                            user = ""
                        one = {}
                        one['参数值'] = "user#" + user
                        one["一级类型"] = systemType
                        one["二级类型"] = appType
                        one["应用名称"] = appName
                        one["节点id"] = nodeId
                        one["参数"] = "database|" + database.attrib.get("id") + ":" + "auth"
                        one["参数说明"] = "id为【"+ database.attrib.get("id") +"】的数据库user"
                        params.append(one)

                        password = auth.attrib.get("password")
                        if password is None:
                            password = ""
                        one = {}
                        one['参数值'] = "password#" + password
                        one["一级类型"] = systemType
                        one["二级类型"] = appType
                        one["应用名称"] = appName
                        one["节点id"] = nodeId
                        one["参数"] = "database|" + database.attrib.get("id") + ":" + "auth"
                        one["参数说明"] = "id为【"+ database.attrib.get("id") +"】的数据库密码"
                        params.append(one)

                    type = database.attrib.get("type")
                    if type is None:
                        type = "mysql"
                    one = {}
                    one['参数值'] = type
                    one["一级类型"] = systemType
                    one["二级类型"] = appType
                    one["应用名称"] = appName
                    one["节点id"] = nodeId
                    one["参数"] = "database|" + database.attrib.get("id") + ":" + "type"
                    one["参数说明"] = "id为【"+ database.attrib.get("id") +"】的数据库类型"
                    params.append(one)

                    enable = database.attrib.get("enable")
                    if enable is None:
                        enable = "true"
                    one = {}
                    one['参数值'] = enable
                    one["一级类型"] = systemType
                    one["二级类型"] = appType
                    one["应用名称"] = appName
                    one["节点id"] = nodeId
                    one["参数"] = "database|" + database.attrib.get("id") + ":" + "enable"
                    one["参数说明"] = "id为【"+ database.attrib.get("id") +"】的数据库是否启用"
                    params.append(one)

                    backup = database.attrib.get("backup")
                    if backup is None:
                        backup = "true"
                    one = {}
                    one['参数值'] = backup
                    one["一级类型"] = systemType
                    one["二级类型"] = appType
                    one["应用名称"] = appName
                    one["节点id"] = nodeId
                    one["参数"] = "database|" + database.attrib.get("id") + ":" + "backup"
                    one["参数说明"] = "id为【"+ database.attrib.get("id") +"】的数据库是否备份"
                    params.append(one)




def write_excel_append(path, sheet_name, dateframe=None):
    # 参数说明: [变量顺序可改变，依次是：sheet页对象，要写入的dataframe对象，从哪一行开始写入]

    writer = pd.ExcelWriter(path, engine='openpyxl', mode='a')  # 用于首次写入还可自动加表头
    workbook = load_workbook(path)  # 打开要写入数据的工作簿
    writer.book = workbook
    rows = dateframe.shape[0]  # 获得行数
    cols = dateframe.shape[1]  # 列数
    # 如果sheet不存在,直接写入，存在，从指定位置写入
    if sheet_name in workbook.sheetnames:
        start_row = workbook[sheet_name].max_row
    else:
        dateframe.to_excel(writer, sheet_name=sheet_name, index=False, header=True)
        writer.save()
        return
    sheet = workbook[sheet_name]  # 打开要编辑的工作表

    if start_row <= 1:
        #  sheet 已存在，直接追加会新增一个同名的sheet页，所以先复制再追加。
        writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)  # 复制已存在的sheet
        dateframe.to_excel(writer, sheet_name=sheet_name, index=False, header=True)
        writer.save()
        return

    for i in range(0, rows):
        for j in range(0, cols):  # value.shape[1]获得列数
            sheet.cell(row=start_row + 1 + i, column=j + 1, value=dateframe.iloc[i][j])
    workbook.save(path)


def r_find_all(root_tag, target='field', type=None):
    """
    遍历根标签，查询目标属性的标签
    :param root_tag: 根标签
    :param target: 需要查找的标签
    :param type: 需要查找的标签属性
    :return: 命中的标签列表
    """
    if root_tag is None or target is None: return []
    try:
        lists = list(root_tag.iter())  # 当前根节点对应的所有子元素包含当前标签
    except:
        lists = []
        print("子节点为空……")

    re = []
    while len(lists) > 0:
        root = lists.pop(0)
        if root.tag == target:
            if type is not None:
                if root.attrib["type"] == type:
                    re.append(root)
            else:
                re.append(root)
    return re


# 将sheet页中的其中几列作为key（k_col_indexs控制列号集合），整行作为value
def sheet2map(path='F:\\test\\test.xlsx', sheet_name='部署包配置页', k_col_indexs=[1,2,3], isConfig=False, cols=5):
    map = {}
    workbook = load_workbook(path)
    worksheet = workbook[sheet_name]
    if len(k_col_indexs) == 0:
        # 默认参数配置页的3、4、5列，对应应用名称、节点id、参数
        k_col_indexs = [3, 4, 5]

    for i in range(1, worksheet.max_row):
        key = ""
        if isConfig and str(worksheet.cell(row=i + 1, column=11).value) != "覆盖":
            continue
        for k_col_index in k_col_indexs:
            tmp = worksheet.cell(row=i + 1, column=k_col_index).value
            if tmp is None:
                tmp = ""
            key = key + "#" + str(tmp).strip()
        val = []
        tmpi = 1
        for item in list(worksheet.rows)[i]:
            if tmpi > cols:
                break
            val.append(item.value)
            tmpi = tmpi + 1
        if key.startswith('##'):
            # map[key] = val 不会报错
            print('\033[4;33m' + '第【{}】行【{}】列为空，已跳过……'.format(i, k_col_indexs[0]) + '\033[0m')
        else:
            if val is None:
                print('\033[4;33m' + '第【{}】行为空'.format(i) + '\033[0m')
            map[key] = val
    print("读取【" + sheet_name + "】sheet页转为json如下：")
    workbook.close()
    print(map)
    return map

def xml2excel(xml_path=None, excel_path=None, curpath=None, map={}):

    try:
        if xml_path == None or excel_path == None:
            return

        with open(xml_path, 'tr', encoding='utf-8') as rf:
            tree = ET.parse(rf)
            root = tree.getroot()
            basic = root.find('basic')
            systemTypes = basic.find('systemType')
            primaryType = basic.find('primaryType')
            if systemTypes != None:
                systemType = systemTypes.text
            else:
                if primaryType != None:
                    systemType = primaryType.text
                else:
                    systemType = '未解析'

            appTypes = basic.find('appType')
            secondaryType = basic.find('secondaryType')
            if appTypes != None:
                appType = appTypes.text
            else:
                if secondaryType != None:
                    appType = secondaryType.text
                else:
                    appType = '未解析'

            appName = basic.find('appName').text
            version = str(basic.find("vrsion").text).strip()
            key = systemType + '#' + appType + "#" + appName + "#"

            val = curpath + "##" +  version
            map[key] = val
            return map
    except:
        print('解析xml出错了……')
        return False


def deal_zip(zip=None, zip_name='', excel_path=None, map={}):
    re = False
    if zip == None or excel_path == None:
        print('excel路径错误，本次跳过……')
        return
    # 打印zip文件中的文件列表
    contains = [x for i, x in enumerate(zip.namelist()) if x.find('deploy.xml') != -1]
    if len(contains) > 0:
        print('当前压缩文件【{}】存在deploy.xml，提取文件。'.format(zip_name))
        xml = zip.extract(contains[0], path=None, pwd=None)
        re = xml2excel(xml, excel_path, zip_name, map)
        os.remove(xml)
        print('当前压缩文件【{}】处理完成！\n'.format(zip_name))
    else:

        for filename in zip.namelist():
            if filename.__contains__('core_sdk'):
                coresdk = zip.extract(filename, path=None, pwd=None)
                sdkzip = zipfile.ZipFile(coresdk, "r")

                contains1 = [x for i, x in enumerate(sdkzip.namelist()) if x.find('deploy.xml') != -1]
                if len(contains1) > 0:
                    xml = sdkzip.extract(contains1[0], path=None, pwd=None)
                    re = xml2excel(xml, excel_path, zip_name,map)
                    os.remove(xml)
                    print('当前压缩文件【{}】处理完成！\n'.format(zip_name))
                else:
                    print('当前压缩文件【{}】未找到deploy.xml,跳过……'.format(zip_name))


                sdkzip.close()
                os.remove(str(coresdk))
                break

    zip.close()
    return re


def main(dirss, excel_path):
    succ = []
    fail = []
    print('开始读取目录：【' + dirss + '】下的压缩文件')
    success_count = 0
    fail_count = 0
    zipmap = {}
    for (root, dirs, files) in os.walk(dirss):
        for f in files:
            curpath = os.path.join(root, f)
            if str(f).endswith('.zip'):
                print('【' + curpath + '】')
                try:
                    z = zipfile.ZipFile(curpath, "r")
                    result = deal_zip(z, curpath, excel_path, zipmap)
                    if result:
                        success_count = success_count + 1
                        succ.append(curpath)
                    else:
                        fail_count = fail_count + 1
                        fail.append(curpath)

                except Exception as e:
                    fail_count = fail_count + 1
                    fail.append(curpath)
                    print('Error:' + str(e.args))
                    print('文件：【' + curpath + '】读取失败，本次跳过……')
                    continue

    print('读取目录：【' + dirss + '】及其子目录下的所有压缩文件结束……')
    print('本次处理压缩文件成功：【{}】个，失败【{}】个'.format(success_count, fail_count))
    print('success: ')
    for s in succ:
        print(s)

    print('fail: ')
    for s in fail:
        print(s)

    print("zipmap:")
    print(zipmap)
    sheetmap = sheet2map("D:\\方案配置文档.xlsx")
    print("sheetmap:")
    print(sheetmap)




def load_conf(path='./conf/conf.ini'):
    print('正在加载配置文件……')
    config = configparser.ConfigParser()
    try:
        config.read(filenames=path, encoding='utf-8')  # utf-8-sig
        print("安装包路径：【{}】".format(config.get("path", "dir")))
        print("excel路径：【{}】".format(config.get("path", "excel_path")))
    except:
        print('加载失败，请检查配置文件conf/conf.ini……')
        print('3秒后自动退出……')
        for i in range(0, 3):
            print(str(3 - i) + '……')
            time.sleep(1)
        sys.exit('end……')
    return {'dir': config.get("path", "dir"), 'excel_path': config.get("path", "excel_path")}


if __name__ == '__main__':

    while True:
        conf = load_conf()
        if os.path.isdir(conf.get('dir')):
            print('安装包路径OK……')
        else:
            print('安装包路径不存在，请修改配置后按 enter……')
            skip = input()
            continue

        if os.path.exists(conf.get('excel_path')) and conf.get('excel_path').endswith('.xlsx'):
            print('excel路径OK……')
            break
        else:
            print('excel路径错误，请修改配置后按 enter……')
            skip = input()
            continue

    print('---start---')
    main(conf.get('dir'), conf.get('excel_path'))
    print('---end---')
    skip = input()
