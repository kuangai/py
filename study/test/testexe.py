import configparser
import os
import sys
import time
import xml.etree.ElementTree as ET
import zipfile
import pandas as pd
from openpyxl import load_workbook
from pandas import DataFrame


def excel2map(path='D:\\test\\test.xlsx', sheet_name='Sheet1', k_col_index=3, v_col_index=4):
    """
    excel2map
    :param k_col_index:
    :param v_col_index:
    :return:
    """
    map = {}
    workbook = load_workbook(path)
    worksheet = workbook[sheet_name]
    for i in range(1, worksheet.max_row):
        key = worksheet.cell(row=i + 1, column=k_col_index).value
        # val = worksheet.cell(row=i, column=v_col_index).value
        val = [item.value for item in list(worksheet.rows)[i]]  # 整行
        if key is None:
            # map[key] = val 不会报错
            print('\033[4;33m' + '第【{}】行【{}】列为空，已跳过……'.format(i, k_col_index) + '\033[0m')
        else:
            if val is None:
                print('\033[4;33m' + '第【{}】行为空'.format(i) + '\033[0m')
            map[key] = val

    print(map)
    return map


def hidden_sheet(path='D:\\test\\test.xlsx', sheet_name='方案名称'):
    workbook = load_workbook(path)  # 打开要写入数据的工作簿
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]  # 打开要编辑的工作表
        sheet.sheet_state = 'hidden'
    workbook.save(path)


def write_excel_node(path, sheet_name, listmap=[]):
    if len(listmap) == 0:
        return True
    try:
        workbook = load_workbook(path)  # 打开要写入数据的工作簿
        sheet = workbook[sheet_name]  # 打开要编辑的工作表
        for j in range(0, len(listmap)):  # value.shape[1]获得列数
            sheet.cell(row=1, column=j + 6, value=listmap[j]['node'])
            sheet.cell(row=2, column=j + 6, value=listmap[j]['user'])
            sheet.cell(row=3, column=j + 6, value=listmap[j]['selected'])
        workbook.save(path)
        workbook.close()
        print('方案名称 sheet 写入成功……')
        return True
    except Exception as e:
        print('方案名称 sheet 写入失败……', e.args)
        if e.args.__contains__('Permission denied'):
            print("Error: 【请关闭待写入的excel】")
        return False


def deal_grid_params(xml,excel_path,col,grid):
    col = []

    packagegrid = DataFrame(columns= col  )  # 生成空的pandas表

    try:
        writer = pd.ExcelWriter(excel_path, engine='openpyxl', mode='a')  # 用于首次写入还可自动加表头
        workbook = load_workbook(excel_path)  # 打开要写入数据的工作簿
        writer.book = workbook

        for k in range(0, len(packagelist)):
            var = packagelist[k]
            s = []
            s.append(var['部署包类型'])
            s.append(var['一级类型'])
            s.append(var['二级类型'])
            s.append(var['应用名称'])
            s.append(var['安装顺序'])
            s.append(var['部署包名称'])
            s.append(var['最低兼容版本'])
            s.append(var['最高兼容版本'])
            packagedf.loc[k] = s

        if sheet_name in workbook.sheetnames:
            workbook.remove(workbook[sheet_name])

        packagedf.to_excel(writer, sheet_name=sheet_name, index=False, header=True)
        writer.save()
        workbook.close()
        print("写入安装包列表sheet写入成功……")
        return True
    except Exception as e:
        print("写入安装包列表sheet写入失败……", e.args)
        if e.args.__contains__('Permission denied'):
            print("Error: 【请关闭待写入的excel】")
        return False
    return []


def write_excel_append(path, sheet_name, dateframe=None):
    # 参数说明: [变量顺序可改变，依次是：sheet页对象，要写入的dataframe对象，从哪一行开始写入]

    try:
        writer = pd.ExcelWriter(path, engine='openpyxl', mode='a')  # 用于首次写入还可自动加表头
        workbook = load_workbook(path)  # 打开要写入数据的工作簿
        writer.book = workbook
        rows = dateframe.shape[0]  # 获得行数
        cols = dateframe.shape[1]  # 列数
        # 如果sheet不存在,直接写入，存在，从指定位置写入
        if sheet_name in workbook.sheetnames:
            start_row = workbook[sheet_name].max_row
        else:
            dateframe.to_excel(writer, sheet_name=sheet_name, index=False, header=True)
            writer.save()
            writer.close()
            return True
        sheet = workbook[sheet_name]  # 打开要编辑的工作表

        if start_row <= 1:
            #  sheet 已存在，直接追加会新增一个同名的sheet页，所以先复制再追加。
            writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)  # 复制已存在的sheet
            dateframe.to_excel(writer, sheet_name=sheet_name, index=False, header=True)
            writer.save()
            writer.close()
            return True

        for i in range(0, rows):
            for j in range(0, cols):  # value.shape[1]获得列数
                sheet.cell(row=start_row + 1 + i, column=j + 1, value=dateframe.iloc[i][j])
        workbook.save(path)
        workbook.close()
        return True
    except Exception as e:
        print("参数配置表sheet 写入失败…… ", e.args)
        if e.args.__contains__('Permission denied'):
            print("Error: 【请关闭待写入的excel】")
        return False


def r_find_all(root_tag, target='field', type=None):
    """
    遍历根标签，查询目标属性的标签
    :param root_tag: 根标签
    :param target: 需要查找的标签
    :param type: 需要查找的标签属性
    :return: 命中的标签列表
    """
    if root_tag is None or target is None: return []
    try:
        lists = list(root_tag.iter())  # 当前根节点对应的所有子元素包含当前标签
    except:
        lists = []
        print("子节点为空……")

    re = []
    while len(lists) > 0:
        root = lists.pop(0)
        if root.tag == target:
            if type is not None:
                if root.attrib["type"] == type:
                    re.append(root)
            else:
                re.append(root)
    return re


def write_excel_package(excel_path=None, sheet_name="安装包列表", packagelist=[]):
    packagedf = DataFrame(columns=('部署包类型', '一级类型', '二级类型', '应用名称', '安装顺序', '部署包名称', '最低兼容版本', '最高兼容版本'))  # 生成空的pandas表

    try:
        writer = pd.ExcelWriter(excel_path, engine='openpyxl', mode='a')  # 用于首次写入还可自动加表头
        workbook = load_workbook(excel_path)  # 打开要写入数据的工作簿
        writer.book = workbook

        for k in range(0, len(packagelist)):
            var = packagelist[k]
            s = []
            s.append(var['部署包类型'])
            s.append(var['一级类型'])
            s.append(var['二级类型'])
            s.append(var['应用名称'])
            s.append(var['安装顺序'])
            s.append(var['部署包名称'])
            s.append(var['最低兼容版本'])
            s.append(var['最高兼容版本'])
            packagedf.loc[k] = s

        if sheet_name in workbook.sheetnames:
            workbook.remove(workbook[sheet_name])

        packagedf.to_excel(writer, sheet_name=sheet_name, index=False, header=True)
        writer.save()
        workbook.close()
        print("写入安装包列表sheet写入成功……")
        return True
    except Exception as e:
        print("写入安装包列表sheet写入失败……", e.args)
        if e.args.__contains__('Permission denied'):
            print("Error: 【请关闭待写入的excel】")
        return False


def xml2excel(xml_path=None, excel_path=None, lists={}, nodemaplist=[], packagelist=[]):
    """
    xml数据以追加的方式转存excel
    :param xml_path:
    :param excel_path:
    :param lists: 从上层穿过来的部分参数，（excel生成方案文档配置页 的数据）
    :param packagelist 每次解析时补充安装的信息到列表，用于最后生成 安装包列表sheet
    :param nodemaplist 每次解析时补充节点的信息到列表，用于最后补入 方案名称sheet
    :return: 处理结果
    """
    support_param_types = ['input', 'select', 'timestamp', 'switch']
    try:
        if xml_path == None or excel_path == None:
            return
        params = []

        with open(xml_path, 'tr', encoding='utf-8') as rf:
            tree = ET.parse(rf)
            root = tree.getroot()
            basic = root.find('basic')
            systemType = lists[0]
            version = basic.find('version')
            appType = lists[1]
            appName = lists[2]

            # 收集部署包sheet需要的数据
            packagemap = {}
            packagemap["部署包类型"] = "组件" if 'inner' == systemType else "应用"
            packagemap["一级类型"] = systemType
            packagemap["二级类型"] = appType
            packagemap["应用名称"] = appName
            packagemap["安装顺序"] = lists[3]
            packagemap["部署包名称"] = str(lists[4]).split('\\')[-1]
            packagemap["最低兼容版本"] = version.text  # todo
            packagemap["最高兼容版本"] = version.text
            packagelist.append(packagemap)

            subSystems = root.find('subSystems')
            systems = subSystems.findall('system')
            for i in range(0, len(systems)):
                sys = systems[i]
                if sys == None:
                    continue
                nodemap = {}

                # 收集方案名称sheet需要的参数
                nodemap['user'] = sys.attrib['id']  # todo 用户
                nodemap['node'] = sys.attrib['id']
                nodemap['selected'] = '√'
                nodemaplist.append(nodemap)


                variables = sys.find('variables')
                if variables == None:
                    continue

                # 单独处理grid参数，由于grid参数标签层级不固定，因此类似递归查询，每个grid单独生成新的sheet页
                grid_fields = r_find_all(sys, target='field', type='grid')
                print('当前节点【{}】对应的grid类型参数共【{}】个', sys.attrib['id'], len(grid_fields))
                for grid in grid_fields:
                    sheetname = grid.attrib.get('name') + grid.attrib.get('label')
                    gridparam = {}
                    gridparam['参数值'] = "grid：" + sheetname
                    gridparam["一级类型"] = systemType
                    gridparam["二级类型"] = appType
                    gridparam["应用名称"] = appName
                    gridparam["节点id"] = sys.attrib['id']
                    gridparam["参数"] = field.attrib.get('name')
                    gridparam["参数说明"] = field.attrib.get('label')

                fields = variables.findall('field')

                for field in fields:

                    if field.text != None \
                            and field.attrib.get('type') is not None \
                            and support_param_types.__contains__(field.attrib.get('type')):
                        one = {}
                        one['参数值'] = field.text
                        one["一级类型"] = systemType
                        one["二级类型"] = appType
                        one["应用名称"] = appName
                        one["节点id"] = sys.attrib['id']
                        one["参数"] = field.attrib.get('name')
                        one["参数说明"] = field.attrib.get('label')
                        # one["参数类型"]
                        params.append(one)

        # print (date)
        paramsdf = DataFrame(
            columns=('一级类型', '二级类型', '应用名称', '节点id', '参数', '参数说明', '参数值', '参数类型', '参数覆盖', '参数新增时间'))  # 生成空的pandas表

        re = False
        sheet_name = '参数配置表'
        try:
            for k in range(0, len(params)):
                var = params[k]
                s = []
                s.append(var['一级类型'])
                s.append(var['二级类型'])
                s.append(var['应用名称'])
                s.append(var['节点id'])
                s.append(var['参数'])
                s.append(var['参数说明'])
                s.append(var['参数值'])
                s.append("")  # todo
                s.append("")
                s.append("")
                paramsdf.loc[k] = s
            print('本次读取参数如下：')
            print(paramsdf)
            print("开始写入参数配置表 sheet……")
            re = write_excel_append(excel_path, sheet_name, paramsdf)
            if re:
                print("写入参数配置表 sheet成功……")
            else:
                print("写入参数配置表 sheet失败……")

        except Exception as e:
            print("写入参数配置表 sheet失败……", e.args)
            if e.args.__contains__('Permission denied'):
                print("Error: 【请关闭待写入的excel】")
            return False
        rf.close()
        print("this time xml2excel execute is fine")
        return re
    except Exception as e:
        print('解析xml出错了……', e.args)
        return False


def deal_zip(zip=None, zip_name='', excel_path=None, lists={}, nodemaplist=[], packagelist=[]):
    re = False
    if zip == None or excel_path == None:
        print('excel路径错误，本次跳过……')
        return
    # 打印zip文件中的文件列表
    contains = [x for i, x in enumerate(zip.namelist()) if x.find('deploy.xml') != -1]
    if len(contains) > 0:
        print('当前压缩文件【{}】存在deploy.xml，提取文件。'.format(zip_name))
        xml = zip.extract(contains[0], path=None, pwd=None)
        re = xml2excel(xml, excel_path, lists, nodemaplist, packagelist)
        os.remove(xml)
        print('当前压缩文件【{}】处理完成！\n'.format(zip_name))
    else:

        for filename in zip.namelist():
            if filename.__contains__('core_sdk'):
                coresdk = zip.extract(filename, path=None, pwd=None)
                sdkzip = zipfile.ZipFile(coresdk, "r")

                contains1 = [x for i, x in enumerate(sdkzip.namelist()) if x.find('deploy.xml') != -1]
                if len(contains1) > 0:
                    xml = sdkzip.extract(contains1[0], path=None, pwd=None)
                    re = xml2excel(xml, excel_path, lists, nodemaplist, packagelist)
                    os.remove(xml)
                    print('当前压缩文件【{}】处理完成！\n'.format(zip_name))
                else:
                    print('当前压缩文件【{}】未找到deploy.xml,跳过……'.format(zip_name))
                    re = False

                sdkzip.close()
                os.remove(str(coresdk))
                break

    zip.close()
    return re


def main(excel_path):
    succ = []
    fail = []
    success_count = 0
    fail_count = 0
    nodemaplist = []
    packagelist = []
    map = excel2map(excel_path, "生成方案文档配置页")
    checks = []
    for f in map:
        ppath = map[f][4]
        if not os.path.exists(ppath):
            checks.append(ppath)
            continue

    if len(checks) > 0:
        print('以下安装包路径填写错误，请检查……')
        print(checks)
        return False

    for f in map:
        curpath = map[f][4]
        print(f, ':', '【' + curpath + '】')
        try:
            z = zipfile.ZipFile(curpath, "r")
            result = deal_zip(z, curpath, excel_path, map[f], nodemaplist, packagelist)
            if result:
                success_count = success_count + 1
                succ.append(curpath)
            else:
                fail_count = fail_count + 1
                fail.append(curpath)

        except Exception as e:
            fail_count = fail_count + 1
            fail.append(curpath)
            print('Error:' + str(e.args))
            print('文件：【' + curpath + '】读取失败，本次跳过……')
            continue

    print("nodemaplist", nodemaplist)
    print("packagelist", packagelist)

    write_excel_node(excel_path, "方案名称", nodemaplist)
    write_excel_package(excel_path, "安装包列表", packagelist)

    print('本次处理压缩文件成功：【{}】个，失败【{}】个'.format(success_count, fail_count))
    print('success: ')
    for s in succ:
        print(s)

    print('fail: ')
    for s in fail:
        print(s)


def load_conf(path='./conf/conf.ini'):
    print('正在加载配置文件……')
    config = configparser.ConfigParser()
    try:
        config.read(filenames=path, encoding='utf-8')  # 搞不定就换 utf-8-sig
        print("安装包路径：【{}】".format(config.get("path", "dir")))
        print("excel路径：【{}】".format(config.get("path", "excel_path")))
    except:
        print('加载失败，请检查配置文件conf/conf.ini……')
        print('3秒后自动退出……')
        for i in range(0, 3):
            print(str(3 - i) + '……')
            time.sleep(1)
        sys.exit('end……')
    return {'dir': config.get("path", "dir"), 'excel_path': config.get("path", "excel_path")}


if __name__ == '__main__':

    while True:
        conf = load_conf()

        if os.path.exists(conf.get('excel_path')) and conf.get('excel_path').endswith('.xlsx'):
            print('excel路径OK……')
            break
        else:
            print('excel路径错误，请修改配置后按 enter……')
            skip = input()
            continue

    print('---start---')
    time1 = time.time()
    main(conf.get('excel_path'))
    time2 = time.time()
    print('---end---spent time: ' + str(int(time2 - time1)) + 's')
    skip = input()
