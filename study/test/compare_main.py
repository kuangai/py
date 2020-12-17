import configparser
import os
import sys
import time
import xml.etree.ElementTree as ET
import zipfile
import pandas as pd
from openpyxl import load_workbook
from pandas import DataFrame
import shutil


def write_excel_append(path, sheet_name, dateframe=None):
    # 参数说明: [变量顺序可改变，依次是：sheet页对象，要写入的dataframe对象，从哪一行开始写入]

    try:
        writer = pd.ExcelWriter(path, engine='openpyxl', mode='a')  # 用于首次写入还可自动加表头
        workbook = load_workbook(path)  # 打开要写入数据的工作簿
        writer.book = workbook
        rows = dateframe.shape[0]  # 获得行数
        cols = dateframe.shape[1]  # 列数
        # 如果sheet不存在,直接写入，存在，从指定位置写入
        if sheet_name in workbook.sheetnames:
            start_row = workbook[sheet_name].max_row
        else:
            dateframe.to_excel(writer, sheet_name=sheet_name, index=False, header=True)
            writer.save()
            writer.close()
            return True
        sheet = workbook[sheet_name]  # 打开要编辑的工作表

        if start_row <= 1:
            #  sheet 已存在，直接追加会新增一个同名的sheet页，所以先复制再追加。
            writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)  # 复制已存在的sheet
            dateframe.to_excel(writer, sheet_name=sheet_name, index=False, header=True)
            writer.save()
            writer.close()
            return True

        for i in range(0, rows):
            for j in range(0, cols):  # value.shape[1]获得列数
                sheet.cell(row=start_row + 1 + i, column=j + 1, value=dateframe.iloc[i][j])
        workbook.save(path)
        workbook.close()
        return True
    except Exception as e:
        print("Error 参数配置表sheet 写入失败…… ", e.args, e.__traceback__.tb_lineno)
        if e.args.__contains__('Permission denied'):
            print("Error: 【请关闭待写入的excel】")
        return False


def sheet2map(path='F:\\test\\test.xlsx', sheet_name='参数配置表', k_col_indexs=[], cols =10):
    map = {}
    workbook = load_workbook(path)
    worksheet = workbook[sheet_name]
    if len(k_col_indexs) == 0:
        # 默认参数配置页的3、4、5列，对应应用名称、节点id、参数
        k_col_indexs = [3, 4, 5]

    for i in range(1, worksheet.max_row):
        key = ""
        for k_col_index in k_col_indexs:
            tmp = worksheet.cell(row=i + 1, column=k_col_index).value
            if tmp is None:
                tmp = ""
            key = key + "#" + str(tmp).strip()
        if str(worksheet.cell(row=i+1, column=5).value).strip().startswith("database") and str(worksheet.cell(row=i+1, column=5).value).strip().endswith("auth"):
            key = key + "#" + str(worksheet.cell(row=i+1, column=7).value).strip().split("#")[0]
        val = []
        tmpi = 1
        for item in list(worksheet.rows)[i]:
            if tmpi > cols:
                break
            val.append(item.value)
            tmpi = tmpi + 1
        if key.startswith('##'):
            # map[key] = val 不会报错
            print('\033[4;33m' + '第【{}】行【{}】列为空，已跳过……'.format(i, k_col_indexs[0]) + '\033[0m')
        else:
            if val is None:
                print('\033[4;33m' + '第【{}】行为空'.format(i) + '\033[0m')
            if key in map.keys():
                print(key+' 存在重复，请检查excel')
            map[key] = val
    print("读取【" + sheet_name + "】sheet页转为json如下：")
    workbook.close()
    print(map)
    return map


def compare(source_path, source_sheet_name, target_path, target_sheet_name):
    source_map = sheet2map(source_path, source_sheet_name, [1, 2, 3, 4, 5], 10)
    target_map = sheet2map(target_path, target_sheet_name, [1, 2, 3, 4, 5], 10)

    workbook = load_workbook(target_path)  # 打开要写入数据的工作簿
    # sheet = workbook[sheet_name]  # 打开要编辑的工作表

    for target_key in target_map.keys():
        target_value = target_map.get(target_key)
        source_value = source_map.get(target_key)
        if source_value is None:
            target_value.append("现场未配置")
            target_value.append(target_value[6])
            target_value.append(None)
        elif str(target_value[6]).strip() == str(source_value[6]).strip():
            target_value.append("一致")
            target_value.append(target_value[6])
            target_value.append(source_value[6])
        else:
            target_value.append("不一致")
            target_value.append(target_value[6])
            target_value.append(source_value[6])

    for source_key in source_map.keys():
        target_value = target_map.get(source_key)
        source_value = source_map.get(source_key)
        if target_value is None:
            source_value.append("包内未读到")
            source_value.append(None)
            source_value.append(source_value[6])
            target_map[source_key] = source_value


    paramsdf = DataFrame(
        columns=('一级类型', '二级类型', '应用名称',
                 '节点id', '参数', '参数说明', '参数值',
                 '参数类型', '参数覆盖', '参数新增时间',
                 "对比结果", "deploy值", "现场值"))  # 生成空的pandas表
    sheet_name = '比对结果'
    for k in range(0, len(target_map)):
        paramsdf.loc[k] = list(target_map.values())[k]
    print("开始写入参数配置表 sheet……")

    workbook = load_workbook(target_path)
    if sheet_name in workbook.sheetnames:
        workbook.remove(workbook[sheet_name])
        workbook.close()

    re = write_excel_append(target_path, sheet_name, paramsdf)
    print("a")


if __name__ == '__main__':
    source_path = "F:\部署包\O45_20201122\筛选.xlsx"
    source_sheet_name = "来自现场"
    target_path = "F:\部署包\O45_20201122\筛选.xlsx"
    target_sheet_name = "来自deploy"
    # try:
    print('---start---')
    time1 = time.time()
    compare(source_path, source_sheet_name, target_path, target_sheet_name)
    time2 = time.time()
    print('---end---spent time: ' + str(int(time2 - time1)) + 's')
    print(input('enter键结束……'))

    # except Exception as e:
    #     print("出现异常..", e.args, e.__traceback__.tb_lineno)

