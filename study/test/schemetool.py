# -*- coding: UTF-8 -*-
import configparser
import json
import os
import sys
import time
import traceback
import xml.etree.ElementTree as ET
import zipfile
import pandas as pd
from openpyxl import load_workbook
from pandas import DataFrame
import shutil
import logging
from logging import handlers
from openpyxl.utils import get_column_letter
import io
from openpyxl.comments import Comment
from time import mktime
from datetime import datetime

reload(sys)
sys.setdefaultencoding('UTF-8')


class Logger(object):
    level_relations = {
        'debug': logging.DEBUG,
        'info': logging.INFO,
        'warning': logging.WARNING,
        'error': logging.ERROR,
        'critical': logging.CRITICAL
    }  # 日志级别关系映射

    def __init__(self, filename, level='info', when='D', backCount=3,
                 fmt='%(asctime)s - %(pathname)s[line:%(lineno)d] - %(levelname)s: %(message)s'):
        output_dir = "log"
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        self.logger = logging.getLogger(filename)
        format_str = logging.Formatter(fmt)  # 设置日志格式
        self.logger.setLevel(self.level_relations.get(level))  # 设置日志级别
        sh = logging.StreamHandler()  # 往屏幕上输出
        sh.setFormatter(format_str)  # 设置屏幕上显示的格式
        th = handlers.TimedRotatingFileHandler(filename=filename, when=when, backupCount=backCount,
                                               encoding='UTF-8')  # 往文件里写入#指定间隔时间自动生成文件的处理器
        # 实例化TimedRotatingFileHandler
        # interval是时间间隔，backupCount是备份文件的个数，如果超过这个个数，就会自动删除，when是间隔的时间单位，单位有以下几种：
        # S 秒
        # M 分
        # H 小时、
        # D 天、
        # W 每星期（interval==0时代表星期一）
        # midnight 每天凌晨
        th.setFormatter(format_str)  # 设置文件里写入的格式
        self.logger.addHandler(sh)  # 把对象加到logger里
        self.logger.addHandler(th)

def is_valid_date(str):
  '''判断是否是一个有效的日期字符串'''
  try:
    strut_time = time.strptime(str, "%Y-%m-%d")
    return {"valid":True,"strut_time":strut_time}
  except:
    try:
        strut_time = time.strptime(str, "%Y-%m-%d %H:%M:%S")
        return {"valid":True,"strut_time":strut_time}
    except:
        return {"valid":False}

#  加批注 该参数仅在上级开关【】开启后生效
def add_comment(parent_map={}, path='F:\\test\\test.xlsx', sheet_name='参数配置表'):
    if len(parent_map.keys()) > 0:
        t1 = time.time()
        log.logger.info("开始在【参数配置表】增加批注……")
        log.logger.debug("开关父级参数映射关系如下：")
        log.logger.debug(parent_map)
        writer = pd.ExcelWriter(path, engine='openpyxl', mode='a')  # 用于首次写入还可自动加表头
        workbook = load_workbook(path)  # 打开要写入数据的工作簿
        writer.book = workbook
        worksheet = workbook[sheet_name]

        for i in range(1, worksheet.max_row):
            systemType = str(worksheet.cell(row=i + 1, column=1).value)
            appType = str(worksheet.cell(row=i + 1, column=2).value)
            appName = str(worksheet.cell(row=i + 1, column=3).value)
            nodeId = str(worksheet.cell(row=i + 1, column=4).value)
            param = str(worksheet.cell(row=i + 1, column=5).value)
            key = systemType + "#" + appType + "#" + appName + "#" + nodeId + "#" + param
            parent_param = parent_map.get(key.decode("utf-8"))
            if parent_param is not None and str(parent_param).strip() != "":
                commonts = "该参数仅在上级开关【{}】开启后生效".format(parent_param)
                comment = Comment(commonts, 'author')
                comment.width = 200
                comment.height = 100
                worksheet["E" + str(i + 1)].comment = comment

        writer.save()
        writer.close()
        t2 = time.time()
        log.logger.info("【参数配置表】增加批注完成……")
        log.logger.info("【参数配置表】增加批注spent time " + str(t2 - t1) + "s")

    return True


def modify_sheet_col_width(path):
    t1 = time.time()
    sheet_list = ['参数配置表','安装包列表','全局参数']

    workbook = load_workbook(path)  # 打开要写入数据的工作簿
    for sheet_name in sheet_list:
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]  # 打开要编辑的工作表
            FullRange = "A1:" + get_column_letter(sheet.max_column) + str(sheet.max_row)
            sheet.auto_filter.ref = FullRange
            sheet.column_dimensions['A'].width = 20
            sheet.column_dimensions['B'].width = 20
            sheet.column_dimensions['C'].width = 35
            sheet.column_dimensions['D'].width = 25
            sheet.column_dimensions['E'].width = 15
            sheet.column_dimensions['F'].width = 35
            sheet.column_dimensions['G'].width = 30
            sheet.column_dimensions['H'].width = 20
            sheet.column_dimensions['I'].width = 10
            sheet.column_dimensions['J'].width = 20
            sheet.column_dimensions['K'].width = 20
            workbook.save(path)
            workbook.close()
    t2 = time.time()
    log.logger.info("调整列宽并设置表头过滤spent time: " + str(t2-t1) + "s")


def load_conf(path='./conf/conf.ini'):
    # log.logger.debug('正在加载配置文件……')
    config = configparser.ConfigParser()
    try:
        config.read(filenames=path, encoding='UTF-8-sig')  # 搞不定就换 UTF-8-sig
        # log.logger.debug(config.get("app", "exclude"))
        # log.logger.debug("配置文件中：生成方案的sheet页顺序：")
        # log.logger.debug(config.get("order", "sheet"))
        # log.logger.debug("配置文件中：日志级别：")
        # log.logger.debug(config.get("log", "level"))

    except Exception as e:
        # log.logger.critical(traceback.format_exc())
        # log.logger.critical('Error: 加载失败，请检查配置文件conf/conf.ini……')
        print("读取配置出错……")
        time.sleep(3)
        sys.exit('end……')
    return {'exclude_app': config.get("app", "exclude"),
            'log_level': config.get("log", "level"),
            'sheet_order': config.get("order", "sheet")}


def conf_xml2excel(xml_path=None, excel_path=None, curpath=None, map={}, package_type=""):
    try:
        if xml_path == None or excel_path == None:
            return

        with io.open(xml_path, 'tr', encoding='UTF-8') as rf:
            tree = ET.parse(rf)
            root = tree.getroot()
            basic = root.find('basic')
            systemTypes = basic.find('systemType')
            primaryType = basic.find('primaryType')
            if systemTypes != None:
                systemType = systemTypes.text.strip()
            else:
                if primaryType != None:
                    systemType = primaryType.text.strip()
                else:
                    systemType = '未解析'

            appTypes = basic.find('appType')
            secondaryType = basic.find('secondaryType')
            if appTypes != None:
                appType = appTypes.text.strip()
            else:
                if secondaryType != None:
                    appType = secondaryType.text.strip()
                else:
                    appType = '未解析'

            appName = basic.find('appName').text.strip()
            version = str(basic.find("version").text).strip()
            key = systemType + '#' + appType + "#" + appName + "#"
            val = curpath + "##" + version
            if map.get(key) is not None and map.get(key) != val:
                app_package_path1 = str(map.get(key))
                if app_package_path1.lower().__contains__(package_type):
                    if val.__contains__(package_type):
                        log.logger.error("配置的部署包路径下应用【" + appName + "】部署包路径中存在多个，路径分别如下，请检查……")
                        log.logger.error(str(map.get(key)).split("##")[0])
                        log.logger.error(curpath)
                        sys.exit("end……")
                    else:
                        # 不处理
                        log.logger.warning(appName + "在路径下部署保存在多个，已检查" + package_type + "版本的使用")
                elif val.lower().__contains__(package_type):
                    log.logger.warning(appName + "在路径下部署保存在多个，已检查" + package_type + "版本的使用")
                    map[key] = val
                else:
                    log.logger.error("配置的部署包路径下应用【" + appName + "】部署包路径中存在多个，路径分别如下，请检查……")
                    log.logger.error(str(map.get(key)).split("##")[0])
                    log.logger.error(curpath)
                    sys.exit("end……")
            else:
                map[key] = val
            return True
    except Exception as e:
        log.logger.critical(traceback.format_exc())
        log.logger.critical('Error: 解析xml出错了……')
        time.sleep(3)
        sys.exit("end……")


# 打印的dataframe不省略
pd.set_option('display.max_rows', 500)
pd.set_option('display.max_columns', 500)
pd.set_option('display.width', 1000)


# 将sheet页中的其中几列作为key（k_col_indexs控制列号集合），整行作为value
def conf_sheet2map(path='F:\\test\\test.xlsx', sheet_name='部署包配置页', k_col_indexs=[1, 2, 3], isConfig=False, cols=5):
    map = {}
    workbook = load_workbook(path)
    worksheet = workbook[sheet_name]
    if len(k_col_indexs) == 0:
        # 默认参数配置页的3、4、5列，对应应用名称、节点id、参数
        k_col_indexs = [3, 4, 5]

    for i in range(1, worksheet.max_row):
        key = ""
        for k_col_index in k_col_indexs:
            tmp = worksheet.cell(row=i + 1, column=k_col_index).value
            if tmp is None:
                tmp = ""
            key = key + str(tmp).strip() + "#"
        val = []
        tmpi = 1
        for item in list(worksheet.rows)[i]:
            if tmpi > cols:
                break
            val.append(item.value)
            tmpi = tmpi + 1
        if key.startswith('##'):
            # map[key] = val 不会报错
            log.logger.warning(sheet_name + '第【{}】行【{}】列为空，已跳过……'.format(i + 1, k_col_indexs[0]))
        else:
            if val is None:
                log.logger.warning(sheet_name + '第【{}】行为空'.format(i))
            map[key] = val
    log.logger.debug("读取【" + sheet_name + "】sheet页转为json如下：")
    workbook.close()
    log.logger.debug(map)
    return map


def conf_json2excel(json_path=None, curpath=None, map={}, package_type=""):
    if json_path is None:
        return True
    with io.open(json_path, 'r', encoding='UTF-8', errors='ignore') as f:
        try:
            info_dict = json.load(f, strict=False)
            if info_dict and len(info_dict) > 0:
                log.logger.debug("本次从文件获取json对象：")
                log.logger.debug(info_dict)
                systemType = info_dict["basic"]["systemType"]
                version = info_dict["basic"]["version"]
                appType = info_dict["basic"]["appType"]
                appName = info_dict["basic"]["appType"]  # 包里的 "appName": "inner-zk-3.4.14",
                systemType = str(systemType).strip()
                appType = str(appType).strip()
                appName = str(appName).strip()
                version = str(version).strip()

                val = curpath + "##" + version
                key = systemType + "#" + appType + "#" + appName + "#"
                if map.get(key) is not None and map.get(key) != val:

                    app_package_path1 = str(map.get(key))
                    if app_package_path1.lower().__contains__(package_type):
                        if val.__contains__(package_type):
                            log.logger.error("配置的部署包路径下应用【" + appName + "】部署包路径中存在多个，路径分别如下，请检查……")
                            log.logger.error(str(map.get(key)).split("##")[0])
                            log.logger.error(curpath)
                            sys.exit("end……")
                        else:
                            # 不处理
                            log.logger.warning(appName + "在路径下部署保存在多个，已检查" + package_type + "版本的使用")
                    elif val.lower().__contains__(package_type):
                        log.logger.warning(appName + "在路径下部署保存在多个，已检查" + package_type + "版本的使用")
                        map[key] = val
                    else:
                        log.logger.error("Error: 配置的部署包路径下应用【" + appName + "】部署包路径中存在多个，路径分别如下，请检查……")
                        log.logger.error(str(map.get(key)).split("##")[0])
                        log.logger.error(curpath)
                        time.sleep(3)
                        sys.exit("end……")
                else:
                    map[key] = val

                return True
            f.close()
        except Exception as e:
            log.logger.critical(traceback.format_exc())
            log.logger.critical("conf_json2excel() 方法读取json文件失败……")
            f.close()
            time.sleep(3)
            sys.exit("end……")
        return False


def conf_deal_zip(zip=None, zip_name='', excel_path=None, map={}, package_type=""):
    re = False
    if zip == None or excel_path == None:
        log.logger.error('excel路径错误，本次跳过……')
        return
    # 打印zip文件中的文件列表
    contains = [x for i, x in enumerate(zip.namelist()) if x.find('deploy.xml') != -1]
    contains_json = [x for i, x in enumerate(zip.namelist()) if x.find('deploy.json') != -1]
    if len(contains) > 0:
        log.logger.info('当前压缩文件【{}】存在deploy.xml，提取文件。'.format(zip_name))
        xml = zip.extract(contains[0], path=None, pwd=None)
        re = conf_xml2excel(xml, excel_path, zip_name, map, package_type)
        os.remove(xml)
        log.logger.info('当前压缩文件【{}】处理完成！\n'.format(zip_name))
    elif len(contains_json) > 0:
        log.logger.info('当前压缩文件【{}】存在deploy.json，提取文件。'.format(zip_name))
        json_path = zip.extract(contains_json[0], path=None, pwd=None)
        re = conf_json2excel(json_path=json_path, curpath=zip_name, map=map, package_type=package_type)
        os.remove(json_path)
        log.logger.info('当前压缩文件【{}】处理完成！\n'.format(zip_name))
    else:

        for filename in zip.namelist():
            if filename.__contains__('core_sdk'):
                coresdk = zip.extract(filename, path=None, pwd=None)
                sdkzip = zipfile.ZipFile(coresdk, "r")

                contains1_json = [x for i, x in enumerate(sdkzip.namelist()) if x.find('deploy.json') != -1]
                contains1 = [x for i, x in enumerate(sdkzip.namelist()) if x.find('deploy.xml') != -1]
                if len(contains1) > 0:
                    xml = sdkzip.extract(contains1[0], path=None, pwd=None)
                    re = conf_xml2excel(xml, excel_path, zip_name, map, package_type)
                    os.remove(xml)
                    log.logger.info('当前压缩文件【{}】处理完成！\n'.format(zip_name))
                elif len(contains1_json) > 0:
                    json_path = sdkzip.extract(contains1_json[0], path=None, pwd=None)
                    re = conf_json2excel(json_path=json_path, curpath=zip_name, map=map, package_type=package_type)
                    os.remove(json_path)
                    log.logger.info('当前压缩文件【{}】处理完成！\n'.format(zip_name))

                else:
                    log.logger.error('当前压缩文件【{}】未找到deploy.xml,跳过……'.format(zip_name))

                sdkzip.close()
                os.remove(str(coresdk))
                break

    zip.close()
    return re


def conf_main(dirss, excel_path, package_type):
    t1 = time.time()
    log.logger.info("开始处理目录，通过目录下部署包更新方案配置……")
    succ = []
    fail = []
    log.logger.info('读取目录：【' + dirss + '】下的压缩文件')
    success_count = 0
    fail_count = 0
    zipmap = {}
    for (root, dirs, files) in os.walk(dirss):
        for f in files:
            # f = f.decode('gbk').encode('utf8')
            curpath = os.path.join(root, f)
            if str(f).endswith('.zip'):
                log.logger.debug('遍历的路径：【' + curpath + '】')
                try:
                    z = zipfile.ZipFile(curpath, "r")
                    result = conf_deal_zip(z, curpath, excel_path, zipmap, package_type)
                    if result:
                        success_count = success_count + 1
                        succ.append(curpath)
                    else:
                        fail_count = fail_count + 1
                        fail.append(curpath)

                except Exception as e:
                    log.logger.critical(traceback.format_exc())
                    fail_count = fail_count + 1
                    fail.append(curpath)
                    log.logger.critical(str(e.args))
                    log.logger.critical('文件：【' + curpath + '】读取失败，程序退出……')
                    time.sleep(3)
                    sys.exit("end……")

    log.logger.info('读取目录：【' + dirss + '】及其子目录下的所有压缩文件结束……')
    log.logger.info('本次读取压缩文件成功：【{}】个'.format(success_count))
    log.logger.info('success: ')
    for s in succ:
        log.logger.info(s)

    sheetmap = conf_sheet2map(excel_path)

    if sheetmap is None or len(sheetmap.keys()) == 0:
        log.logger.error("Error: excel配置中部署包配置页为空，不再处理，程序退出……")
        sys.exit("end……")
    else:
        log.logger.debug("部署包配置页中配置如下:")
        log.logger.debug(sheetmap)

    if zipmap is None or len(zipmap.keys()) == 0:
        log.logger.error("Error: 配置的部署包路径中部署包为空，不再处理，程序退出……")
        sys.exit("end……")
    else:
        log.logger.debug("配置路径下：部署包路径##版本号读取成功如下:")
        log.logger.debug(zipmap)

    log.logger.info("开始收集部署包中配置的路径……")
    for key in sheetmap.keys():
        if zipmap.get(key.decode("utf-8")) is None and str(zipmap.get(key.decode("utf-8"))).strip() == "":
            app_name = key.split("#")[2]
            log.logger.error("Error: 应用【" + app_name + "】部署包不存在，请检查部署包路径……")
            time.sleep(3)
            sys.exit("end……")

        cols_list = sheetmap.get(key)
        cols_list.append(str(zipmap.get(key.decode("utf-8"))).split("##")[0])
    log.logger.info("生成方案配置详细信息如下……")
    for key in sheetmap.keys():
        log.logger.info("[" + str(sheetmap.get(key)[0]) + ", " + str(sheetmap.get(key)[1]) + ", " + str(
            sheetmap.get(key)[2]) + ", " +
                        str(sheetmap.get(key)[3]) + ", " + str(sheetmap.get(key)[4]) + "] ")

    t2 = time.time()
    log.logger.info("自动匹配安装包路径完成 spent time：" + str(t2-t1) + "s……")
    return sheetmap


def deal_json_params(excel_path="", package_list=[], info_dict=None, nodemaplist=[], lists={}, cover_map={}):
    if info_dict is None:
        return True
    t1 = time.time()
    systemType = info_dict["basic"]["systemType"]
    version = info_dict["basic"]["version"]
    appType = info_dict["basic"]["appType"]
    appName = info_dict["basic"]["appType"]  # 包里的 "appName": "inner-zk-3.4.14",
    systems = info_dict["system"]
    # 收集部署包sheet需要的数据
    packagemap = {}
    packagemap["部署包类型"] = "组件" if 'inner' == systemType else "应用"
    packagemap["一级类型"] = systemType
    packagemap["二级类型"] = appType
    packagemap["应用名称"] = appName
    packagemap["安装顺序"] = lists[3]
    packagemap["部署包名称"] = str(lists[4]).split('\\')[-1]
    packagemap["最低兼容版本"] = version  # todo
    packagemap["最高兼容版本"] = version

    package_list.append(packagemap)
    # 收集方案名称sheet需要的参数

    if len(systems) > 0:
        for system in systems:
            variables = system["variable"]
            if len(variables) > 0:
                for variable in variables:
                    if system["id"] is not None and \
                            variable["name"] is not None \
                            and variable["name"] == 'user' \
                            and variable["value"] is not None:
                        nodemap = {}
                        nodemap['user'] = variable["value"]
                        nodemap['node'] = system["id"]
                        nodemaplist.append(nodemap)
    params = []

    paramsdf = DataFrame(
        columns=('一级类型', '二级类型', '应用名称', '节点id', '参数', '参数说明', '参数值', '参数类型', '参数覆盖', '参数新增时间'))  # 生成空的pandas表

    modify_parameter_config(cover_map, params, appName)

    re = False
    sheet_name = '参数配置表'

    try:
        for k in range(0, len(params)):
            var = params[k]
            s = []
            s.append(str(var['一级类型']).strip())
            s.append(str(var['二级类型']).strip())
            s.append(str(var['应用名称']).strip())
            s.append('' if var['节点id'] is None else str(var['节点id']).strip())
            s.append(var['参数'])
            s.append(var['参数说明'])
            s.append(var['参数值'])
            s.append(var['参数类型'])
            s.append("")

            datestr = var['参数新增时间']
            if datestr is not None and str(datestr).strip() != '':
                datestr = str(datestr.date())
            s.append(datestr)

            paramsdf.loc[k] = s
        log.logger.debug('本次读取常规参数如下：')
        log.logger.debug(paramsdf)
        log.logger.debug("开始写入参数配置表 sheet……")
        re = write_excel_append(excel_path, sheet_name, paramsdf)
        if re:
            t2 = time.time()
            log.logger.info("【" + appName + "】写入参数配置表 sheet成功 spent time: " + str(t2-t1) + "s……")
        else:
            log.logger.info("写入参数配置表 sheet失败……")
        return re
    except Exception as e:
        log.logger.critical(traceback.format_exc())
        log.logger.critical("Error 写入参数配置表 sheet失败……")
        if e.args.__contains__('Permission denied'):
            log.logger.critical("Error: 【请关闭待写入的excel】")
        time.sleep(3)
        sys.exit("end……")


def json2excel(excel_path="", package_list=[], json_path=None, nodemaplist=[], lists={}, cover_map={}):
    """
    处理zk的deploy.json
    暂不解析参数，只把zk的部署包信息和节点用户存入安装包列表sheet和方案名称sheet
    :return:
    """
    if json_path is None:
        return
    with io.open(json_path, 'r', encoding='UTF-8', errors='ignore') as f:
        try:
            info_dict = json.load(f, strict=False)
            if info_dict and len(info_dict) > 0:
                log.logger.debug("本次从文件获取json对象：")
                log.logger.debug(info_dict)
                deal_json_params(excel_path, package_list, info_dict, nodemaplist, lists=lists, cover_map=cover_map)
                return True
            else:
                log.logger.error("json参数为空，不再处理……")
                return False
            f.close()
        except Exception as e:
            log.logger.critical(traceback.format_exc())
            log.logger.critical('Error  json 参数处理失败……')
            f.close()
            time.sleep(3)
            sys.exit("end……")
        return False


def deal_inner_field_self(field, params, systemType, appType, appName, nodeId, filter_map):
    # 加入参数
    one = {}
    if field.text is None:
        text = ""
        if field.attrib.get("default") is not None:
            text = field.attrib.get("default")
    else:
        text = field.text.strip()
        if text == "":
            if field.attrib.get("default") is not None:
                text = field.attrib.get("default")

    one['参数值'] = text
    one["一级类型"] = systemType
    one["二级类型"] = appType
    one["应用名称"] = appName
    one["节点id"] = nodeId
    one["参数"] = field.attrib.get('name')
    one["参数说明"] = field.attrib.get('label')

    zgfieldtype = field.attrib.get('zgfieldtype')
    if zgfieldtype is None or str(zgfieldtype) == '':
        zgfieldtype = "默认"
    one["参数类型"] = zgfieldtype
    zgfiledtime = field.attrib.get('zgfiledtime')
    if zgfiledtime is None:
        zgfiledtime = ""
    one["参数新增时间"] = zgfiledtime
    isfilter = filter_map.get(str(appName + "#" + one["节点id"] + "#" + one["参数"] + "#"))
    if (isfilter is None or isfilter is not True):
        params.append(one)


def deal_inner_field(parent_map, field1, support_param_types, params, systemType, appType, appName, nodeId, filter_map):
    listfield = []
    field_map = {}
    field_map["field"] = field1
    field_map["parent"] = ""
    listfield.append(field_map)
    while len(listfield) > 0:
        tmpfieldmap = listfield.pop(0)
        tmpfield = tmpfieldmap.get("field")
        if tmpfield is None:
            continue

        parent_params = tmpfieldmap.get("parent")
        next_parent_param = ''
        if tmpfield.attrib.get("type") is not None and tmpfield.attrib.get('type') == "switchForm":
            next_parent_param = tmpfield.attrib.get('name')
            deal_inner_field_self(tmpfield, params, systemType, appType, appName, nodeId, filter_map)
            if parent_params != "":
                parent_map[systemType + "#" + appType + "#" + appName + "#" + nodeId + "#" + tmpfield.attrib.get(
                    'name') + "#"] = parent_params

        innerfields = tmpfield.findall("field")

        if len(innerfields) > 0:
            for field in innerfields:

                if field.attrib.get("type") is not None \
                        and (support_param_types.__contains__(field.attrib.get('type'))):
                    # 加入参数
                    one = {}
                    if field.text is None:
                        text = ""
                        if field.attrib.get("default") is not None:
                            text = field.attrib.get("default")
                    else:
                        text = field.text.strip()
                        if text == "":
                            if field.attrib.get("default") is not None:
                                text = field.attrib.get("default")

                    one['参数值'] = text
                    one["一级类型"] = systemType
                    one["二级类型"] = appType
                    one["应用名称"] = appName
                    one["节点id"] = nodeId
                    one["参数"] = field.attrib.get('name')
                    one["参数说明"] = field.attrib.get('label')
                    if next_parent_param != "":
                        parent_map[systemType + "#" + appType + "#" + appName + "#" + nodeId + "#" + field.attrib.get(
                            'name')] = next_parent_param

                    zgfieldtype = field.attrib.get('zgfieldtype')
                    if zgfieldtype is None or str(zgfieldtype) == '':
                        zgfieldtype = "默认"
                    one["参数类型"] = zgfieldtype
                    zgfiledtime = field.attrib.get('zgfiledtime')
                    if zgfiledtime is None:
                        zgfiledtime = ""
                    one["参数新增时间"] = zgfiledtime
                    isfilter = filter_map.get(str(appName + "#" + one["节点id"] + "#" + one["参数"] + "#"))
                    if (isfilter is None or isfilter is not True):
                        params.append(one)

                else:
                    if field.attrib.get("type") is not None and field.attrib.get("type") == 'grid':
                        pass
                    # 其他类型，判断是否有嵌套
                    else:
                        if len(field.findall("field")) > 0:
                            field_map = {}
                            field_map["field"] = field
                            field_map["parent"] = next_parent_param
                            listfield.append(field_map)

        else:
            continue


def get_real_sheet_name(excel_path, sheet_name):
    writer = pd.ExcelWriter(excel_path, engine='openpyxl', mode='a')  # 用于首次写入还可自动加表头
    workbook = load_workbook(excel_path)  # 打开要写入数据的工作簿
    writer.book = workbook
    tmpi = 2
    tmp_sheet_name = sheet_name
    while tmp_sheet_name in workbook.sheetnames:
        tmp_sheet_name = sheet_name + str(tmpi)
        tmpi = tmpi + 1
    workbook.close()
    return tmp_sheet_name


def deal_database_param(databases=None, params=None, systemType=None, appType=None, appName=None, nodeId=None,
                        filter_map={}):
    if databases is None or params is None:
        return
    if databases is not None:
        database_list = databases.findall("database")
        if database_list is not None and len(database_list) > 0:
            for database in database_list:
                if database.attrib.get("id") is not None:
                    is_match = (database.find("matchers") is not None) or (database.find("matcher") is not None)
                    auth = database.find("auth")
                    if auth is not None:
                        user = auth.attrib.get("user")
                        if user is None:
                            user = ""
                        one = {}
                        one['参数值'] = "user#" + user
                        one["一级类型"] = systemType
                        one["二级类型"] = appType
                        one["应用名称"] = appName
                        one["节点id"] = nodeId
                        one["参数"] = "database|" + database.attrib.get("id") + ":" + "auth"
                        one["参数说明"] = "id为【" + database.attrib.get("id") + "】的数据库user,参数值#后配置用户名"
                        one["参数类型"] = "数据库"
                        one["参数新增时间"] = ''
                        isfilter = filter_map.get(str(appName + "#" + nodeId + "#" + one["参数"] + "#"))
                        if bool(1 - is_match) and (isfilter is None or isfilter is not True):
                            params.append(one)

                        # 't2_grid_channel_mgr_servers服务治理集群'
                        password = auth.attrib.get("password")
                        if password is None:
                            password = ""
                        one = {}
                        one['参数值'] = "password#" + password
                        one["一级类型"] = systemType
                        one["二级类型"] = appType
                        one["应用名称"] = appName
                        one["节点id"] = nodeId
                        one["参数"] = "database|" + database.attrib.get("id") + ":" + "auth"
                        one["参数说明"] = "id为【" + database.attrib.get("id") + "】的数据库密码,参数值#后配置密码"
                        one["参数类型"] = "数据库"
                        one["参数新增时间"] = ''
                        isfilter = filter_map.get(str(appName + "#" + nodeId + "#" + one["参数"] + "#"))
                        if bool(1 - is_match) and (isfilter is None or isfilter is not True):
                            params.append(one)

                    type = database.attrib.get("type")
                    if type is None:
                        type = "mysql"
                    one = {}
                    one['参数值'] = type
                    one["一级类型"] = systemType
                    one["二级类型"] = appType
                    one["应用名称"] = appName
                    one["节点id"] = nodeId
                    one["参数"] = "database|" + database.attrib.get("id") + ":" + "type"
                    one["参数说明"] = "id为【" + database.attrib.get("id") + "】的数据库类型"
                    one["参数类型"] = "数据库"
                    one["参数新增时间"] = ''
                    isfilter = filter_map.get(str(appName + "#" + nodeId + "#" + one["参数"] + "#"))
                    if bool(1 - is_match) and (isfilter is None or isfilter is not True):
                        params.append(one)

                    enable = database.attrib.get("enable")
                    enabled = database.attrib.get("enabled")

                    if enable is None:
                        if enabled is None:
                            enables = "true"
                        else:
                            enables = enabled
                    else:
                        enables = enable
                    one = {}
                    one['参数值'] = enables
                    one["一级类型"] = systemType
                    one["二级类型"] = appType
                    one["应用名称"] = appName
                    one["节点id"] = nodeId
                    one["参数"] = "database|" + database.attrib.get("id") + ":" + "enable"
                    one["参数说明"] = "id为【" + database.attrib.get("id") + "】的数据库是否启用"
                    one["参数类型"] = "数据库"
                    one["参数新增时间"] = ''
                    isfilter = filter_map.get(str(appName + "#" + nodeId + "#" + one["参数"] + "#"))
                    if isfilter is None or isfilter is not True:
                        params.append(one)

                    backup = database.attrib.get("backup")
                    if backup is None:
                        backup = "true"
                    one = {}
                    one['参数值'] = backup
                    one["一级类型"] = systemType
                    one["二级类型"] = appType
                    one["应用名称"] = appName
                    one["节点id"] = nodeId
                    one["参数"] = "database|" + database.attrib.get("id") + ":" + "backup"
                    one["参数说明"] = "id为【" + database.attrib.get("id") + "】的数据库是否备份"
                    one["参数类型"] = "数据库"
                    one["参数新增时间"] = ''
                    isfilter = filter_map.get(str(appName + "#" + nodeId + "#" + one["参数"] + "#"))
                    if isfilter is None or isfilter is not True:
                        params.append(one)

                    user = database.attrib.get("user")
                    if user is not None:
                        one = {}
                        one['参数值'] = user
                        one["一级类型"] = systemType
                        one["二级类型"] = appType
                        one["应用名称"] = appName
                        one["节点id"] = nodeId
                        one["参数"] = "databases|selectedDatabases|user:" + database.attrib.get("id")
                        one["参数说明"] = "id为【" + database.attrib.get("id") + "】的数据库user"
                        one["参数类型"] = "数据库"
                        one["参数新增时间"] = ''
                        isfilter = filter_map.get(str(appName + "#" + nodeId + "#" + one["参数"] + "#"))
                        if enables.__ne__("false") and (isfilter is None or isfilter is not True):
                            params.append(one)

                    host = database.attrib.get("host")
                    if host is None:
                        host = ""
                    one = {}
                    one['参数值'] = host
                    one["一级类型"] = systemType
                    one["二级类型"] = appType
                    one["应用名称"] = appName
                    one["节点id"] = nodeId
                    one["参数"] = "database|" + database.attrib.get("id") + ":" + "host"
                    one["参数说明"] = "id为【" + database.attrib.get("id") + "】的数据库host"
                    one["参数类型"] = "数据库"
                    one["参数新增时间"] = ''
                    isfilter = filter_map.get(str(appName + "#" + nodeId + "#" + one["参数"] + "#"))
                    if bool(1 - is_match) and (isfilter is None or isfilter is not True):
                        params.append(one)

                    port = database.attrib.get("port")
                    if port is None:
                        port = ""
                    one = {}
                    one['参数值'] = port
                    one["一级类型"] = systemType
                    one["二级类型"] = appType
                    one["应用名称"] = appName
                    one["节点id"] = nodeId
                    one["参数"] = "database|" + database.attrib.get("id") + ":" + "port"
                    one["参数说明"] = "id为【" + database.attrib.get("id") + "】的数据库port"
                    one["参数类型"] = "数据库"
                    one["参数新增时间"] = ''
                    isfilter = filter_map.get(str(appName + "#" + nodeId + "#" + one["参数"] + "#"))
                    if bool(1 - is_match) and (isfilter is None or isfilter is not True):
                        params.append(one)

                    # 服务名
                    database_str = database.attrib.get("database")
                    if database_str is not None and str(database_str).strip() != "":
                        one = {}
                        one['参数值'] = database_str
                        one["一级类型"] = systemType
                        one["二级类型"] = appType
                        one["应用名称"] = appName
                        one["节点id"] = nodeId
                        one["参数"] = "database|" + database.attrib.get("id") + ":" + "database"
                        one["参数说明"] = "id为【" + database.attrib.get("id") + "】的数据库名或服务名"
                        one["参数类型"] = "数据库"
                        one["参数新增时间"] = ''
                        isfilter = filter_map.get(str(appName + "#" + nodeId + "#" + one["参数"] + "#"))
                        if bool(1 - is_match) and (isfilter is None or isfilter is not True):
                            params.append(one)


def hidden_sheet(path='D:\\test\\test.xlsx'):
    t1 = time.time()
    hidden_list = ['全局变量配置页', '部署包配置页', '默认参数配置页']
    workbook = load_workbook(path)  # 打开要写入数据的工作簿
    for sheet_name in hidden_list:
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]  # 打开要编辑的工作表
            sheet.sheet_state = 'hidden'
    workbook.save(path)
    workbook.close()
    t2 = time.time()
    log.logger.info("隐藏sheet  spent time: " + str(t2 - t1) + "s")


def show_sheet(path='D:\\test\\test.xlsx'):
    t1 = time.time()
    show_list = ['参数配置表']
    workbook = load_workbook(path)  # 打开要写入数据的工作簿
    for sheet_name in show_list:
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]  # 打开要编辑的工作表
            sheet.sheet_state = 'visible'
    workbook.save(path)
    workbook.close()
    t2 = time.time()
    log.logger.info("显示sheet  spent time: " + str(t2 - t1) + "s")


def write_excel_node(path, sheet_name, listmap=[]):
    t1 = time.time()
    if len(listmap) == 0:
        return True
    try:
        workbook = load_workbook(path)  # 打开要写入数据的工作簿
        sheet = workbook[sheet_name]  # 打开要编辑的工作表
        for j in range(0, len(listmap)):  # value.shape[1]获得列数
            nodecell = sheet.cell(row=1, column=j + 6).value
            usercell = sheet.cell(row=2, column=j + 6).value
            if usercell is not None and str(usercell) != "":
                continue
            if nodecell is None or str(nodecell) == '':
                break
            else:
                for k in range(0, len(listmap)):
                    if str(nodecell).lower().replace(".", "") == str(listmap[k]['node']).lower().replace(".", ""):
                        sheet.cell(row=2, column=j + 6, value=listmap[k]['user'])

            #  sheet.cell(row=3, column=j + 6, value=listmap[j]['selected'])
        workbook.save(path)
        workbook.close()
        t2 = time.time()
        log.logger.info("【方案名称】 sheet写入成功spent time: " + str(t2 - t1) + "s……")
        return True
    except Exception as e:
        log.logger.critical(traceback.format_exc())
        log.logger.critical('Error 方案名称 sheet 写入失败……')
        if e.args.__contains__('Permission denied'):
            log.logger.critical("Error: 【请关闭待写入的excel】")
        time.sleep(3)
        sys.exit("end……")



def deal_node_params(parent_map={}, node=None, params=None, systemType=None, appType=None, appName=None, nodeId=None,
                     support_param_types=None, filter_map={}):
    if node is None or nodeId is None:
        return
    variables = node.find('variables')
    if variables is None:
        return
    fields = variables.findall('field')
    for field in fields:
        if field is not None \
                and field.attrib.get('type') is not None \
                and support_param_types.__contains__(field.attrib.get('type')):
            one = {}

            if field.text is None:
                text = ""
            else:
                text = field.text.strip()

            param_val = text
            if param_val is None or param_val.strip() == '':
                param_val = field.attrib.get('default')

            if param_val is None:
                continue

            if field.attrib.get('name') != "grid":
                param_val = "[all|" + str(param_val) + "]"

            one['参数值'] = param_val
            one["一级类型"] = systemType
            one["二级类型"] = appType
            one["应用名称"] = appName
            one["节点id"] = nodeId
            one["参数"] = field.attrib.get('name')
            one["参数说明"] = field.attrib.get('label')
            zgfieldtype = field.attrib.get('zgfieldtype')
            if zgfieldtype is None or str(zgfieldtype) == '':
                zgfieldtype = "默认"
            one["参数类型"] = zgfieldtype
            zgfiledtime = field.attrib.get('zgfiledtime')
            if zgfiledtime is None:
                zgfiledtime = ""
            one["参数新增时间"] = zgfiledtime
            isfilter = filter_map.get(str(appName + "#" + nodeId + "#" + one["参数"] + "#"))
            if isfilter is None or isfilter is not True:
                params.append(one)
        else:
            if field is not None and field.attrib.get("type") != 'grid':
                deal_inner_field(parent_map, field, support_param_types, params, systemType, appType, appName, nodeId,
                                 filter_map)


def deal_grid_params(excel_path=None, grid_tag=None, sheet_name=None, deal_flag=True):
    """
    grid参数生成sheet
    :param sheet_name:
    :param excel_path:
    :param grid_tag: grid标签对象
    :return: 处理结果 True or False
    """
    if deal_flag is False:
        return True
    if excel_path is None or grid_tag is None or sheet_name is None:
        return False

    grid_name = grid_tag.attrib.get('label')
    if grid_name is None:
        grid_name = grid_tag.attrib.get('name')
    fields = grid_tag.findall('field')
    if len(fields) == 0:
        log.logger.error('grid：【{}】 缺少field字段,请检查deploy.xml，本次不再处理……'.format(grid_name))
        return False

    col = []
    for field in fields:
        col.append(field.attrib.get('name'))
    grid_df = DataFrame(columns=tuple(col))  # 生成空的pandas表

    data_str = grid_tag.text
    log.logger.debug("grid内容如下：")
    log.logger.debug(data_str.strip())
    log.logger.debug("grid标签如下：")
    for field in fields:
        log.logger.debug(field.attrib.get("name"))
    if data_str is None or data_str.strip() == '':
        tmp = []
        for field in fields:
            str = field.text
            if str is None:
                str = ""

            tmp.append(str.strip())
        grid_df.loc[0] = tmp
    else:
        tmp_arr = data_str.strip().split(';')
        if len(tmp_arr) == 0:
            return
        for k in range(0, len(tmp_arr)):
            if (len(tmp_arr[k].split(',')) == len(fields)):
                grid_df.loc[k] = tmp_arr[k].split(',')
    log.logger.debug(grid_name + '  参数如下：')
    log.logger.debug(grid_df)

    try:
        re = write_excel_append(excel_path, sheet_name, grid_df)
        if re:
            log.logger.info(grid_name + " grid sheet写入成功……")
            log.logger.info('\n')
            return True
        else:
            log.logger.info(grid_name + " grid sheet写入失败……")
            return False
    except Exception as e:
        log.logger.critical(traceback.format_exc())
        log.logger.critical("Error" + grid_name + " grid sheet写入失败……")
        if e.args.__contains__('Permission denied'):
            log.logger.critical("Error: 【请关闭待写入的excel】")
        time.sleep(3)
        sys.exit("end……")


def write_excel_append(path, sheet_name, dateframe=None):
    # 参数说明: [变量顺序可改变，依次是：sheet页对象，要写入的dataframe对象，从哪一行开始写入]

    try:
        writer = pd.ExcelWriter(path, engine='openpyxl', mode='a')  # 用于首次写入还可自动加表头
        workbook = load_workbook(path)  # 打开要写入数据的工作簿
        writer.book = workbook
        rows = dateframe.shape[0]  # 获得行数
        cols = dateframe.shape[1]  # 列数
        # 如果sheet不存在,直接写入，存在，从指定位置写入
        if sheet_name in workbook.sheetnames:
            start_row = workbook[sheet_name].max_row
        else:
            dateframe.to_excel(writer, sheet_name=sheet_name.decode("utf-8"), index=False, header=True)
            writer.save()
            writer.close()
            return True
        sheet = workbook[sheet_name]  # 打开要编辑的工作表

        if start_row <= 1:
            #  sheet 已存在，直接追加会新增一个同名的sheet页，所以先复制再追加。
            writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)  # 复制已存在的sheet
            dateframe.to_excel(writer, sheet_name=sheet_name.decode("utf-8"), index=False, header=True)
            writer.save()
            writer.close()
            return True

        for i in range(0, rows):
            for j in range(0, cols):  # value.shape[1]获得列数
                sheet.cell(row=start_row + 1 + i, column=j + 1, value=dateframe.iloc[i][j])
        workbook.save(path)
        workbook.close()
        return True
    except Exception as e:
        log.logger.critical(traceback.format_exc())
        log.logger.critical("Error 参数配置表sheet 写入失败…… ")
        if e.args.__contains__('Permission denied'):
            log.logger.critical("Error: 【请关闭待写入的excel】")
        time.sleep(3)
        sys.exit("end……")


def r_find_all(root_tag, target='field', type=None):
    """
    遍历根标签，查询目标属性的标签
    :param root_tag: 根标签
    :param target: 需要查找的标签
    :param type: 需要查找的标签属性
    :return: 命中的标签列表
    """
    if root_tag is None or target is None: return []
    try:
        lists = list(root_tag.iter())  # 当前根节点对应的所有子元素包含当前标签
    except Exception as e:
        log.logger.error(traceback.format_exc())
        lists = []
        log.logger.error("子节点为空……")

    re = []
    while len(lists) > 0:
        root = lists.pop(0)
        if root.tag == target:
            if type is not None:
                if root.attrib["type"] == type:
                    re.append(root)
            else:
                re.append(root)
    return re


def write_excel_package(excel_path=None, sheet_name="安装包列表", packagelist=[]):
    t1 = time.time()
    packagedf = DataFrame(columns=('部署包类型', '一级类型', '二级类型', '应用名称', '安装顺序', '部署包名称', '最低兼容版本', '最高兼容版本'))  # 生成空的pandas表

    packagelist = sorted(packagelist, key=lambda e: str(e.__getitem__('安装顺序')) + e.__getitem__('应用名称'))

    try:
        writer = pd.ExcelWriter(excel_path, engine='openpyxl', mode='a')  # 用于首次写入还可自动加表头
        workbook = load_workbook(excel_path)  # 打开要写入数据的工作簿
        writer.book = workbook

        for k in range(0, len(packagelist)):
            var = packagelist[k]
            s = []
            s.append(var['部署包类型'])
            s.append(var['一级类型'])
            s.append(var['二级类型'])
            s.append(var['应用名称'])
            s.append(var['安装顺序'])
            s.append(var['部署包名称'])
            s.append(var['最低兼容版本'])
            s.append(var['最高兼容版本'])
            packagedf.loc[k] = s

        if sheet_name in workbook.sheetnames:
            workbook.remove(workbook[sheet_name])

        packagedf.to_excel(writer, sheet_name=sheet_name.decode("utf-8"), index=False, header=True)
        writer.save()
        workbook.close()
        t2 = time.time()
        log.logger.info("【安装包列表】 sheet写入成功spent time: " + str(t2 - t1) + "s……")
        return True
    except Exception as e:
        log.logger.critical(traceback.format_exc())
        log.logger.critical("安装包列表sheet写入失败……")
        if e.args.__contains__('Permission denied'):
            log.logger.critical("Error: 【请关闭待写入的excel】")
        time.sleep(3)
        sys.exit("end……")


def xml2excel(parent_map={}, cover_map={}, xml_path=None, excel_path=None, lists={}, nodemaplist=[], packagelist=[],
              filter_map={},
              exclude_app=''):
    """
    xml数据以追加的方式转存excel
    :param xml_path:
    :param excel_path:
    :param lists: 从上层穿过来的部分参数，（excel部署包配置页 的数据）
    :param packagelist 每次解析时补充安装的信息到列表，用于最后生成 安装包列表sheet
    :param nodemaplist 每次解析时补充节点的信息到列表，用于最后补入 方案名称sheet
    :return: 处理结果
    """
    t1 = time.time()
    exclude_app_list = exclude_app.split(";")

    support_param_types = ['password', 'input', 'select', 'timestamp', 'switch', 'complexSelect']
    try:
        if xml_path == None or excel_path == None:
            return
        params = []
        key = ""

        with io.open(xml_path, 'tr', encoding='UTF-8') as rf:
            tree = ET.parse(rf)
            root = tree.getroot()
            basic = root.find('basic')
            systemType = lists[0]
            version = str(basic.find('version').text).strip()
            appType = lists[1]
            appName = basic.find('appName').text
            appName = appName.strip()
            systemType = str(systemType).strip()
            appType = str(appType).strip()
            key = systemType + '#' + appType + '#' + appName
            deal_flag = not exclude_app_list.__contains__(key)

            # 收集部署包sheet需要的数据
            packagemap = {}
            packagemap["部署包类型"] = "组件" if 'inner' == systemType else "应用"
            packagemap["一级类型"] = systemType
            packagemap["二级类型"] = appType
            packagemap["应用名称"] = appName
            packagemap["安装顺序"] = lists[3]
            packagemap["部署包名称"] = str(lists[4]).split('\\')[-1]
            packagemap["最低兼容版本"] = version  # todo
            packagemap["最高兼容版本"] = version
            packagelist.append(packagemap)

            # 全局参数
            global_config = root.find('globalConfig')

            # 数据库参数
            databases = global_config.find('databases')
            deal_database_param(databases, params, systemType, appType, appName, "", filter_map)

            # 常规参数
            global_variables = global_config.find('variables')
            if global_variables is not None:
                # grid
                global_grid_fields = r_find_all(global_variables, target='field', type='grid')
                if len(global_grid_fields) > 0:
                    log.logger.info(appName + ' 全局参数中对应的grid类型参数共【{}】个'.format(len(global_grid_fields)))
                    for grid in global_grid_fields:
                        isfilter = filter_map.get(str(appName + "#" + "#" + grid.attrib.get('name') + "#"))
                        if isfilter:
                            continue
                        sheet_name = grid.attrib.get('label')
                        sheet_name = get_real_sheet_name(excel_path, sheet_name)  # TODO
                        grid_param = {'参数值': "grid：" + sheet_name, "一级类型": systemType, "二级类型": appType, "应用名称": appName,
                                      "节点id": '', "参数": grid.attrib.get('name'),
                                      "参数说明": grid.attrib.get('label')}

                        zgfieldtype = grid.attrib.get('zgfieldtype')
                        if zgfieldtype is None or str(zgfieldtype) == '':
                            zgfieldtype = "默认"
                        grid_param["参数类型"] = zgfieldtype
                        zgfiledtime = grid.attrib.get('zgfiledtime')
                        if zgfiledtime is None:
                            zgfiledtime = ""
                        grid_param["参数新增时间"] = zgfiledtime

                        if isfilter is None or isfilter is not True:
                            params.append(grid_param)
                            deal_grid_params(excel_path, grid, sheet_name, deal_flag)
                # 常规参数
                fields = global_variables.findall('field')
                for field in fields:
                    if field is not None \
                            and field.attrib.get('type') is not None \
                            and support_param_types.__contains__(field.attrib.get('type')):
                        one = {}
                        if field.text is None:
                            text = ""
                        else:
                            text = field.text.strip()
                        one['参数值'] = text
                        one["一级类型"] = systemType
                        one["二级类型"] = appType
                        one["应用名称"] = appName
                        one["节点id"] = ''
                        one["参数"] = field.attrib.get('name')
                        one["参数说明"] = field.attrib.get('label')

                        zgfieldtype = field.attrib.get('zgfieldtype')
                        if zgfieldtype is None or str(zgfieldtype) == '':
                            zgfieldtype = "默认"
                        one["参数类型"] = zgfieldtype
                        zgfiledtime = field.attrib.get('zgfiledtime')
                        if zgfiledtime is None:
                            zgfiledtime = ""
                        one["参数新增时间"] = zgfiledtime
                        isfilter = filter_map.get(str(appName + "#" + one["节点id"] + "#" + one["参数"] + "#"))
                        if isfilter is None or isfilter is not True:
                            params.append(one)
                    else:
                        if field is not None and field.attrib.get("type") != 'grid':
                            deal_inner_field(parent_map, field, support_param_types, params, systemType, appType,
                                             appName, "",
                                             filter_map)
            # 节点参数
            subSystems = root.find('subSystems')
            systems = subSystems.findall('system')
            for i in range(0, len(systems)):
                system = systems[i]
                if sys is None:
                    continue
                nodemap = {}

                # 收集方案名称sheet需要的参数
                nodemap['user'] = ""  # todo 用户
                nodemap['node'] = system.attrib.get('id')
                # nodemap['selected'] = '√'

                # 数据库参数
                databases = system.find('databases')
                deal_database_param(databases, params, systemType, appType, appName, system.attrib.get('id'),
                                    filter_map)

                # 常规参数
                variables = system.find('variables')
                if variables is None:
                    continue

                # 单独处理grid参数，由于grid参数标签层级不固定，因此类似递归查询，每个grid单独生成新的sheet页
                grid_fields = r_find_all(system, target='field', type='grid')
                if len(grid_fields) > 0:
                    log.logger.info(
                        appName + ' 当前节点【{}】对应的grid类型参数共【{}】个'.format(system.attrib.get('id'), len(grid_fields)))
                    time1 = time.time()

                    for grid in grid_fields:
                        isfilter = filter_map.get(
                            str(appName + "#" + system.attrib.get('id') + "#" + grid.attrib.get('name') + "#"))
                        if isfilter:
                            continue
                        sheet_name = grid.attrib.get('label')
                        sheet_name = get_real_sheet_name(excel_path, sheet_name)  # TODO
                        grid_param = {'参数值': "grid：" + sheet_name, "一级类型": systemType, "二级类型": appType, "应用名称": appName,
                                      "节点id": system.attrib.get('id'), "参数": grid.attrib.get('name'),
                                      "参数说明": grid.attrib.get('label')}

                        zgfieldtype = grid.attrib.get('zgfieldtype')
                        if zgfieldtype is None or str(zgfieldtype) == '':
                            zgfieldtype = "默认"
                        grid_param["参数类型"] = zgfieldtype
                        zgfiledtime = grid.attrib.get('zgfiledtime')
                        if zgfiledtime is None:
                            zgfiledtime = ""
                        grid_param["参数新增时间"] = zgfiledtime

                        if isfilter is None or isfilter is not True:
                            params.append(grid_param)
                            deal_grid_params(excel_path, grid, sheet_name, deal_flag)
                    time2 = time.time()
                    log.logger.info("【"+ appName + "】处理grid 参数 spent time："+ str(time2-time1) + "s")
                # 私有节点参数
                node = system.find('node')

                fields = variables.findall('field')
                for field in fields:

                    if field.attrib.get("name") is not None and field.attrib.get(
                            "name") == 'user' and field.text is not None:
                        nodemap['user'] = field.text.strip()

                    if field is not None \
                            and field.attrib.get('type') is not None \
                            and support_param_types.__contains__(field.attrib.get('type')):
                        one = {}
                        if field.text is None:
                            text = ""
                        else:
                            text = field.text.strip()

                        param_val = text
                        if param_val is None or param_val.strip() == '':
                            param_val = field.attrib.get('default')
                        one['参数值'] = param_val
                        one["一级类型"] = systemType
                        one["二级类型"] = appType
                        one["应用名称"] = appName
                        one["节点id"] = system.attrib.get('id')
                        one["参数"] = field.attrib.get('name')
                        one["参数说明"] = field.attrib.get('label')
                        zgfieldtype = field.attrib.get('zgfieldtype')
                        if zgfieldtype is None or str(zgfieldtype) == '':
                            zgfieldtype = "默认"
                        one["参数类型"] = zgfieldtype
                        zgfiledtime = field.attrib.get('zgfiledtime')
                        if zgfiledtime is None:
                            zgfiledtime = ""
                        one["参数新增时间"] = zgfiledtime
                        isfilter = filter_map.get(str(appName + "#" + one["节点id"] + "#" + one["参数"] + "#"))
                        if isfilter is None or isfilter is not True:
                            params.append(one)
                    else:
                        if field is not None and field.attrib.get("type") != 'grid':
                            deal_inner_field(parent_map, field, support_param_types, params, systemType, appType,
                                             appName,
                                             system.attrib.get('id'),
                                             filter_map)

                if node is not None and system.attrib.get('id') is not None:
                    deal_node_params(parent_map, node, params, systemType, appType, appName, system.attrib.get('id'),
                                     support_param_types, filter_map)

                nodemaplist.append(nodemap)

        paramsdf = DataFrame(
            columns=('一级类型', '二级类型', '应用名称', '节点id', '参数', '参数说明', '参数值', '参数类型', '参数覆盖', '参数新增时间'))  # 生成空的pandas表

        # 处理覆盖参数
        if exclude_app_list.__contains__(key):
            params = []
        modify_parameter_config(cover_map, params, appName)

        re = False
        sheet_name = '参数配置表'

        try:
            for k in range(0, len(params)):
                var = params[k]
                if var["参数值"] is None or str(var["参数值"]).strip() == "":
                    continue
                s = []
                s.append(str(var['一级类型']).strip())
                s.append(str(var['二级类型']).strip())
                s.append(str(var['应用名称']).strip())
                s.append('' if var['节点id'] is None else str(var['节点id']).strip())
                s.append(var['参数'])
                s.append(var['参数说明'])
                s.append(var['参数值'])
                s.append(var['参数类型'])
                s.append("")
                datestr = var['参数新增时间']
                if datestr is not None and str(datestr).strip() != '' and str(datestr.__class__) == '<type \'datetime.datetime\'>':
                    datestr = str(datestr.date())
                else:
                    if datestr is not None and str(datestr).strip() != '':
                        # 校验并转换
                        re = is_valid_date(datestr)
                        if re["valid"] is True:
                            datestr = str(datetime.fromtimestamp(mktime(re["strut_time"])).date())

                s.append(datestr)
                paramsdf.loc[k] = s
            log.logger.debug('本次读取常规参数如下：')
            log.logger.debug(paramsdf)
            log.logger.info("开始写入参数配置表 sheet……")
            re = write_excel_append(excel_path, sheet_name, paramsdf)
            if re:
                t2 = time.time()
                log.logger.info("【" + appName + "】写入参数配置表 sheet成功 spent time: " + str(t2 - t1) + "s……")
            else:
                log.logger.error("写入参数配置表 sheet失败……")

        except Exception as e:
            log.logger.critical(traceback.format_exc())
            if e.args.__contains__('Permission denied'):
                log.logger.critical("Error: 【请关闭待写入的excel】")
            time.sleep(3)
            sys.exit("end……")

        rf.close()
        log.logger.debug("this time xml2excel execute is fine")
        return re
    except Exception as e:
        log.logger.critical(traceback.format_exc())
        log.logger.critical('解析xml出错了……')
        time.sleep(3)
        sys.exit("end……")


def deal_zip(parent_map={}, zip=None, zip_name='', excel_path=None, lists={}, nodemaplist=[], packagelist=[],
             exclude_app='',
             filter_map={}, cover_map={}):
    re = False
    if zip == None or excel_path == None:
        log.logger.error('excel路径错误，本次跳过……')
        return
    # 打印zip文件中的文件列表
    contains = [x for i, x in enumerate(zip.namelist()) if x.find('deploy.xml') != -1]
    contains_json = [x for i, x in enumerate(zip.namelist()) if x.find('deploy.json') != -1]

    if len(contains) > 0:
        log.logger.info('当前压缩文件【{}】存在deploy.xml，提取文件。'.format(zip_name))
        xml = zip.extract(contains[0], path=None, pwd=None)
        re = xml2excel(parent_map, cover_map, xml, excel_path, lists, nodemaplist, packagelist, filter_map,
                       exclude_app)
        os.remove(xml)
        log.logger.info('当前压缩文件【{}】处理完成！\n'.format(zip_name))
    elif len(contains_json) > 0:
        log.logger.info('当前压缩文件【{}】存在deploy.json，提取文件。'.format(zip_name))
        json_path = zip.extract(contains_json[0], path=None, pwd=None)
        re = json2excel(excel_path=excel_path, package_list=packagelist, json_path=json_path, nodemaplist=nodemaplist,
                        lists=lists,
                        cover_map=cover_map)
        os.remove(json_path)
        log.logger.info('当前压缩文件【{}】处理完成！\n'.format(zip_name))
    else:
        for filename in zip.namelist():
            if filename.__contains__('core_sdk'):
                coresdk = zip.extract(filename, path=None, pwd=None)
                sdkzip = zipfile.ZipFile(coresdk, "r")

                contains1 = [x for i, x in enumerate(sdkzip.namelist()) if x.find('deploy.xml') != -1]
                contains1_json = [x for i, x in enumerate(sdkzip.namelist()) if x.find('deploy.json') != -1]
                if len(contains1) > 0:
                    xml = sdkzip.extract(contains1[0], path=None, pwd=None)
                    re = xml2excel(parent_map, cover_map, xml, excel_path, lists, nodemaplist, packagelist,
                                   filter_map, exclude_app)
                    os.remove(xml)
                    log.logger.info('当前压缩文件【{}】处理完成！\n'.format(zip_name))
                elif len(contains1_json) > 0:
                    json_path = sdkzip.extract(contains1_json[0], path=None, pwd=None)
                    re = json2excel(excel_path=excel_path, package_list=packagelist, json_path=json_path,
                                    nodemaplist=nodemaplist, lists=lists,
                                    cover_map=cover_map)
                    os.remove(json_path)
                    log.logger.info('当前压缩文件【{}】处理完成！\n'.format(zip_name))
                else:
                    log.logger.error('当前压缩文件【{}】未找到deploy.xml,跳过……'.format(zip_name))
                    re = False

                sdkzip.close()
                os.remove(str(coresdk))
                break

    zip.close()
    return re


# sheet页复制，未包含样式复制
def copySheet(old_sheet_name="全局变量配置页", new_sheet_name="全局变量配置页copy", path="F:\\test\\test.xlsx"):
    workbook = load_workbook(path)
    old_sheet = workbook[old_sheet_name]
    if new_sheet_name in workbook.sheetnames:
        workbook.remove(workbook[new_sheet_name])
    new_sheet = workbook.create_sheet(new_sheet_name)
    for row in old_sheet:
        for cell in row:
            new_sheet[cell.coordinate].value = cell.value
    workbook.save(path)
    workbook.close()


def create_global_var_sheet(path="F:\\test\\test.xlsx"):
    t1 = time.time()
    try:
        copySheet(u"全局变量配置页", u"全局参数", path)
        log.logger.debug("复制全局变量sheet页成功……")
        t2 = time.time()
        log.logger.info("复制【全局变量】sheet页成功……spent time: " + str(t2 - t1) + "s……")
    except Exception as e:
        log.logger.critical(traceback.format_exc())
        log.logger.critical("Error: 复制全局变量sheet页失败……")
        time.sleep(3)
        sys.exit("end……")


# 将sheet页中的其中几列作为key（k_col_indexs控制列号集合），整行作为value
def sheet2map(path='F:\\test\\test.xlsx', sheet_name='参数配置表', k_col_indexs=[], isConfig=False, cols=10):
    map = {}
    workbook = load_workbook(path)
    worksheet = workbook[sheet_name]
    if len(k_col_indexs) == 0:
        # 默认参数配置页的3、4、5列，对应应用名称、节点id、参数
        k_col_indexs = [3, 4, 5]

    for i in range(1, worksheet.max_row):
        key = ""
        if isConfig and str(worksheet.cell(row=i + 1, column=11).value) != "覆盖":
            continue
        for k_col_index in k_col_indexs:
            tmp = worksheet.cell(row=i + 1, column=k_col_index).value
            if tmp is None:
                tmp = ""
            key = key + "#" + str(tmp).strip()
        if key.__contains__(":auth"):
            key = key + "#" + str(worksheet.cell(row=i + 1, column=7).value).split("#")[0]
        val = []
        tmpi = 1
        for item in list(worksheet.rows)[i]:
            if tmpi > cols:
                break
            val.append(item.value)
            tmpi = tmpi + 1
        if key.startswith('##'):
            # map[key] = val 不会报错
            log.logger.error(sheet_name + '第【{}】行【{}】列为空，已跳过……'.format(i + 1, k_col_indexs[0]))
        else:
            if val is None:
                log.logger.error(sheet_name + '第【{}】行为空'.format(i))
            map[key] = val
    log.logger.debug("sheet2map 读取【" + sheet_name + "】sheet页转为json如下：")
    workbook.close()
    log.logger.debug(map)
    return map


# 每个应用的参数放在一起 map<app_name,params_map>  params_map<应用名#节点id#参数#，整行的值>
def appname2paramsmap(path='F:\\test\\test.xlsx', sheet_name='参数配置表', k_col_indexs=[], isConfig=False, cols=10):
    t1 = time.time()
    appname2params = {}
    workbook = load_workbook(path)
    worksheet = workbook[sheet_name]
    if len(k_col_indexs) == 0:
        # 默认参数配置页的3、4、5列，对应应用名称、节点id、参数
        k_col_indexs = [3, 4, 5]

    for i in range(1, worksheet.max_row):
        key = ""
        if isConfig and str(worksheet.cell(row=i + 1, column=11).value) != "覆盖":
            continue
        appName = worksheet.cell(row=i + 1, column=3).value
        params = appname2params.get(appName)
        if params is None:
            params = {}
            appname2params[appName] = params

        for k_col_index in k_col_indexs:
            tmp = worksheet.cell(row=i + 1, column=k_col_index).value
            if tmp is None:
                tmp = ""
            key = key + tmp.strip() + "#"

        if key.__contains__(":auth"):
            key = key + str(worksheet.cell(row=i + 1, column=7).value).split("#")[0] + "#"
        val = []
        tmpi = 1
        for item in list(worksheet.rows)[i]:
            if tmpi > cols:
                break
            val.append(item.value)
            tmpi = tmpi + 1

        if key.startswith('#'):  # 应用名为空跳过
            # map[key] = val 不会报错
            log.logger.debug(sheet_name + '第【{}】行【{}】列应用名为空，已跳过……'.format(i, k_col_indexs[0]))
        else:
            if val is None:
                log.logger.debug(sheet_name + '第【{}】行为空'.format(i))
            params[key] = val

    log.logger.debug("appname2params 读取【" + sheet_name + "】sheet页转为json如下：")
    workbook.close()
    log.logger.debug(appname2params)
    t2 = time.time()
    log.logger.info("读取cover_map完成 spent time：" + str(t2 - t1) + "s……")
    return appname2params


# 将sheet页中的其中几列作为key（k_col_indexs控制列号集合），整行作为value
def sheet2set(path='F:\\test\\test.xlsx', sheet_name='参数配置表', k_col_indexs=[]):
    t1 = time.time()
    map = {}
    workbook = load_workbook(path)
    worksheet = workbook[sheet_name]
    if len(k_col_indexs) == 0:
        # 默认参数配置页的3、4、5列，对应应用名称、节点id、参数
        k_col_indexs = [3, 4, 5]

    for i in range(1, worksheet.max_row):
        key = ""
        if str(worksheet.cell(row=i + 1, column=11).value) == "过滤":
            for k_col_index in k_col_indexs:
                tmp = worksheet.cell(row=i + 1, column=k_col_index).value
                if tmp is None:
                    tmp = ""
                key = key + str(tmp).strip() + "#"
            map[key] = True

    log.logger.debug("读取【" + sheet_name + "】sheet页转为set如下：")
    workbook.close()
    log.logger.debug(map)
    t2 = time.time()
    log.logger.info("读取filter_map完成 spent time：" + str(t2 - t1) + "s……")
    return map


def convert_params(param={}, value_list=[]):
    param["一级类型"] = value_list[0]
    param["二级类型"] = value_list[1]
    param["应用名称"] = value_list[2]
    param["节点id"] = value_list[3]
    param["参数"] = value_list[4]
    param["参数说明"] = value_list[5]
    param['参数值'] = value_list[6]
    param["参数类型"] = value_list[7]
    param["参数覆盖"] = value_list[8]
    param["参数新增时间"] = value_list[9]


def modify_parameter_config(cover_map=None, params=[], app_name=""):
    if cover_map is None or cover_map.get(app_name) is None:
        return
    app_param_map = cover_map.get(app_name)
    if app_param_map is None:
        return

    params_keys = {}
    if len(params) > 0:
        for param in params:
            key = param["应用名称"] + "#" + param["节点id"] + "#" + param["参数"] + "#"
            if param["参数"].__contains__(":auth"):
                key = key + str(param["参数值"]).split("#")[0] + "#"

            params_keys[key] = True
            if app_param_map.get(key) is not None:
                value_list = app_param_map.get(key)  # 覆盖掉旧参数
                convert_params(param, value_list)

    for k in app_param_map.keys():
        if params_keys.get(k.decode("utf-8")) is None:
            param = {}
            value_list = app_param_map.get(k)  # 覆盖掉旧参数
            convert_params(param, value_list)
            params.append(param)  # 新增


def check_default_parameter_config(excel_path):
    t1 = time.time()
    map = sheet2map(excel_path, "默认参数配置页", [], False, 11)
    is_exception = False
    log.logger.info("检查默认参数配置页sheet...")
    for line_num in map:
        judge_field = str(map[line_num][10]).strip()
        if judge_field != "过滤" and judge_field != "覆盖":
            log.logger.error("应用名称：" + str(map[line_num][2]).strip()
                             + " 节点id：" + str(map[line_num][3]).strip()
                             + " 参数：" + str(map[line_num][4]).strip()
                             + " 所在行的 过滤/覆盖 列配置有误")
            is_exception = True
    t2 = time.time()
    log.logger.info("检查【默认参数配置页】sheet完成 spent time：" + str(t2 - t1) + "s……")
    return is_exception


# 获取该sheet所在位置 从0开始
def get_current_order(sheet_name, workbook):
    sheet_names = workbook.sheetnames
    for i in range(0, len(sheet_names)):
        if sheet_names[i] == sheet_name:
            return i


def adjustSheetOrder(excel_path):
    t1 = time.time()
    workbook = load_workbook(excel_path)
    conf = load_conf()
    sheet_order = conf.get("sheet_order").split(";")
    sheet_names = workbook.sheetnames
    for i in range(0, len(sheet_order)):
        tmp = len(sheet_order) - i - 1
        if sheet_order[tmp] in sheet_names:
            current = -get_current_order(sheet_order[tmp], workbook)
            workbook.move_sheet(sheet_order[tmp], current)
        else:
            log.logger.error("sheet页‘" + sheet_order[tmp] + "’调整位置失败..")
    workbook.save(excel_path)
    workbook.close()
    t2 = time.time()
    log.logger.info("调整sheet顺序spent time: " + str(t2 - t1) + "s")


def main(excel_path, exclude_app, dirs, new_excel_path, package_type):
    excel_path_arr = excel_path.split(".xlsx")
    excel_path = new_excel_path
    old_path = excel_path_arr[0] + ".xlsx"
    if os.path.exists(excel_path):
        try:
            os.remove(excel_path)

        except Exception as e:
            log.logger.critical(traceback.format_exc())
            log.logger.critical("请关闭打开的excel文件后重试……")
            time.sleep(3)
            sys.exit("end……")
    shutil.copy(old_path, excel_path)
    succ = []
    fail = []
    success_count = 0
    fail_count = 0
    nodemaplist = []
    packagelist = []
    map = conf_main(dirs, excel_path, package_type)

    checks = []
    for f in map:
        ppath = map[f][4]
        if not os.path.exists(ppath.decode("utf-8")):
            checks.append(f.split("#")[2])
            continue

    if len(checks) > 0:
        log.logger.info('以下安装包未找到对应的路径，本次跳过……')
        for path in checks:
            log.logger.info('应用：【' + path + '】安装包未找到，本次跳过……')
        if len(map.keys()) == len(checks):
            log.logger.critical('所有安装包未找到对应的路径，程序退出·……')
            return False
        # return False

    if check_default_parameter_config(excel_path):
        log.logger.critical('请先修改默认参数配置页中的异常配置...')
        return False

    # sheet_name = "参数配置表"
    # workbook = load_workbook(excel_path)
    # if sheet_name in workbook.sheetnames:
    #     workbook.remove(workbook[sheet_name])
    #     workbook.save(excel_path)
    #     workbook.close()

    filter_map = sheet2set(excel_path, "默认参数配置页", [3, 4, 5])
    cover_map = appname2paramsmap(excel_path, "默认参数配置页", [3, 4, 5], True, 10)
    parent_map = {}  # 记录父级参数增加批注使用
    t1 = time.time()
    for f in map:
        curpath = map[f][4]
        if not os.path.exists(curpath.decode("utf-8")):
            continue
        log.logger.info('正在处理：【' + curpath + '】')
        try:
            z = zipfile.ZipFile(curpath.decode("utf-8"), "r")
            result = deal_zip(parent_map, z, curpath, excel_path, map[f], nodemaplist, packagelist, exclude_app,
                              filter_map,
                              cover_map)
            if result:
                success_count = success_count + 1
                succ.append(curpath)
            else:
                fail_count = fail_count + 1
                fail.append(curpath)

        except Exception as e:
            log.logger.critical(traceback.format_exc())
            fail_count = fail_count + 1
            fail.append(curpath)
            log.logger.critical('文件：【' + curpath + '】读取失败，本次跳过……')
            time.sleep(3)
            sys.exit("end……")
    t2 = time.time()
    log.logger.info("参数处理完成 spent time： " + str(t2-t1) + "s")
    log.logger.debug("nodemaplist: ")
    log.logger.debug(nodemaplist)
    log.logger.debug("packagelist: ")
    log.logger.debug(packagelist)

    write_excel_node(excel_path, "方案名称", nodemaplist)
    write_excel_package(excel_path, "安装包列表", packagelist)
    create_global_var_sheet(excel_path)

    log.logger.info('本次处理压缩文件成功：【{}】个，失败【{}】个'.format(success_count, fail_count))
    if success_count > 0:
        log.logger.info('success: ')
        for s in succ:
            log.logger.info(s)

    hidden_sheet(excel_path)
    show_sheet(excel_path)
    adjustSheetOrder(excel_path)
    modify_sheet_col_width(excel_path)
    add_comment(parent_map, excel_path, "参数配置表")


if __name__ == '__main__':
    conf = load_conf()
    log = Logger('log\schemetool.log', level=conf.get('log_level'))
    exclude_app = ""
    excel_path = ""
    new_excel_path = ""
    dirs = ""

    log.logger.info("---------------------------------------------------------")
    log.logger.info("-----------------------   begin   -----------------------")
    log.logger.info("---------------------------------------------------------")
    for i in range(0, len(sys.argv)):
        log.logger.info(sys.argv[i].decode('gbk').encode('utf8'))
    if len(sys.argv) >= 4 and sys.argv[1].endswith('.xlsx'):
        excel_path = sys.argv[1].decode('gbk').encode('utf8').decode("utf-8")
        dirs = sys.argv[2].decode('gbk').encode('utf8').decode("utf-8")
        new_excel_path = sys.argv[3].decode('gbk').encode('utf8').decode("utf-8")

        package_type = "oracle"
        if len(sys.argv) >= 5:
            tmp_type = sys.argv[4]
            if tmp_type is not None and tmp_type.strip() != '':
                if tmp_type.lower() == 'mysql' or tmp_type.lower() == 'oracle':
                    package_type = tmp_type
                else:
                    log.logger.critical('命令行参数不正确，Usage: ' + sys.argv[
                        0] + ' 配置方案全路径(.xlsx结尾)' + ' 部署包所在目录' + ' 生成方案全路径(.xlsx结尾)' + ' [部署包数据库类型(mysql或oracle)，本参数为可选参数，不填写默认为oracle]')
                    time.sleep(3)
                    sys.exit('end……')

    else:
        log.logger.critical('命令行参数不正确，Usage: ' + sys.argv[
            0] + ' 配置方案全路径(.xlsx结尾)' + ' 部署包所在目录' + ' 生成方案全路径(.xlsx结尾)' + ' [部署包数据库类型(mysql或oracle)，本参数为可选参数，不填写默认为oracle]')
        time.sleep(3)
        sys.exit('end……')

    exclude_app = conf.get('exclude_app')
    if os.path.exists(excel_path) and excel_path.endswith('.xlsx'):
        log.logger.info('配置excel路径OK……')
    else:
        log.logger.critical('第一个参数有误：配置文档不存在' + excel_path)
        time.sleep(3)
        sys.exit('end……')

    if os.path.exists(dirs):
        log.logger.info('部署包路径OK……')
    else:
        log.logger.critical('第二个参数有误：部署包路径不存在' + dirs)
        time.sleep(3)
        sys.exit('end……')

    p, f = os.path.split(new_excel_path)
    if os.path.exists(p) and f.endswith('.xlsx'):
        log.logger.info('生成方案excel路径OK……')
    else:
        log.logger.critical('第三个参数有误：生成方案excel' + new_excel_path)
        time.sleep(3)
        sys.exit('end……')

    log.logger.info('---start---')
    time1 = time.time()
    main(excel_path, exclude_app, dirs, new_excel_path, package_type)
    time2 = time.time()
    log.logger.info("---------------------------------------------------------")
    log.logger.info("------------------------   end   ------------------------")
    log.logger.info("---------------------------------------------------------")
    log.logger.info('--------------------- spent time: ' + str(int(time2 - time1)) + ' s  -----------------')
    log.logger.info("---------------------------------------------------------")
