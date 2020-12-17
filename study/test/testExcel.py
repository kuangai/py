# coding=utf-8
import json
import sys

import pandas as pd

from openpyxl import load_workbook
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.comments import Comment

reload(sys)
sys.setdefaultencoding('UTF-8')

def write_excel_append(path, sheet_name, dateframe=None):
    # 参数说明: [变量顺序可改变，依次是：sheet页对象，要写入的dataframe对象，从哪一行开始写入]

    writer = pd.ExcelWriter(path, engine='openpyxl', mode='a')  # 用于首次写入还可自动加表头
    workbook = load_workbook(path)  # 打开要写入数据的工作簿
    writer.book = workbook  # 激活工作薄
    rows = dateframe.shape[0]  # 获得行数
    cols = dateframe.shape[1]  # 列数
    # 如果sheet不存在,直接写入，存在，从指定位置写入
    if sheet_name in workbook.sheetnames:
        start_row = workbook[sheet_name].max_row
    else:
        dateframe.to_excel(writer, sheet_name=sheet_name, index=False, header=True)
        writer.save()
        return
    sheet = workbook[sheet_name]  # 打开要编辑的工作表

    if start_row <= 1:
        #  sheet 已存在，直接追加会新增一个同名的sheet页，所以先复制再追加。
        writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)  # 复制已存在的sheet
        dateframe.to_excel(writer, sheet_name=sheet_name, index=False, header=True)
        writer.save()
        return

    for i in range(0, rows):
        for j in range(0, cols):  # value.shape[1]获得列数
            sheet.cell(row=start_row + 1 + i, column=j + 1, value=dateframe.iloc[i][j])
    workbook.save(path)


def hidden_sheet(path='D:\\test\\test.xlsx', sheet_name='服务信息'):
    workbook = load_workbook(path)  # 打开要写入数据的工作簿
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]  # 打开要编辑的工作表
        sheet.sheet_state = 'hidden'
    workbook.save(path)


def read_excel(path='D:\\test\\test.xlsx', sheet_name='Sheet1'):
    workbook = load_workbook(path)

    worksheet = workbook.get_sheet_by_name(sheet_name)

    row3 = [item.value for item in list(worksheet.rows)[2]]

    print('第3行值', row3)

    col3 = [item.value for item in list(worksheet.columns)[2]]

    print('第3行值', col3)

    cell_2_3 = worksheet.cell(row=2, column=3).value

    print('第2行第3列值', cell_2_3)

    max_row = worksheet.max_row

    print('最大行', max_row)


def excel_to_json(excel_file, json_file=''):
    wb = load_workbook(excel_file)  # 读取excel文件
    excel_data = {}  # 定义字典excel_data存储每个表的数据{表名:数据}
    for sheet in wb.sheetnames:
        result = []  # 定义列表result存储所有读取数据
        for rows in wb[sheet]:  # 获取表的每一行数据
            tmp = []  # 定义列表tmp存储每一行的数据
            for cell in rows:  # 遍历一行每个单元格的数据
                tmp.append(cell.value)
            result.append(tmp)
        excel_data[sheet] = result
    print(excel_data)
    # 覆盖写入json文件
    # with open(json_file, mode='w', encoding='utf-8') as jf:
    #  json.dump(excel_data, jf, indent=2, sort_keys=True, ensure_ascii=False)


def auto_filter():
    workbook = openpyxl.load_workbook('test.xlsx')
    worksheet = workbook['参数配置表']
    FullRange = "A1:" + get_column_letter(worksheet.max_column) + str(worksheet.max_row)
    worksheet.auto_filter.ref = FullRange
    workbook.save('test.xlsx')


def set_console_print_width():
    pd.set_option('display.max_rows', 500)
    pd.set_option('display.max_columns', 500)
    pd.set_option('display.width', 1000)

def add_excel_comment(file = "D:\\11.xlsx", sheet_name="部署包配置页", commonts="这是批注的内容"):
    writer = pd.ExcelWriter(file, engine='openpyxl', mode='a')  # 用于首次写入还可自动加表头
    workbook = load_workbook(file)  # 打开要写入数据的工作簿
    writer.book = workbook
    # wb.get_sheet_by_name("部署包配置页")
    print(workbook.sheetnames)
    ws = workbook[sheet_name]
    # ws = wb.active  当前使用的sheet,一般都是第一页
    comment = Comment(commonts, 'author')
    # 设置批注框的宽,高
    comment.width = 200
    comment.height = 100
    # 在指定的单元格上面设置批注
    ws['C16'].comment = comment
    writer.save()
    writer.close()

if __name__ == '__main__':
    add_excel_comment()