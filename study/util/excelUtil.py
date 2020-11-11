# coding=utf-8

import pandas as pd

from openpyxl import load_workbook


# 向已存在的sheet追加数据，sheet不存在则创建后追加
def write_excel_append(path, sheet_name, dateframe=None):
    # 参数说明: [变量顺序可改变，依次是：sheet页对象，要写入的dataframe对象，从哪一行开始写入]

    writer = pd.ExcelWriter(path, engine='openpyxl', mode='a')  # 用于首次写入还可自动加表头
    workbook = load_workbook(path)  # 打开要写入数据的工作簿
    writer.book = workbook
    rows = dateframe.shape[0]  # 获得行数
    cols = dateframe.shape[1]  # 列数
    # 如果sheet不存在,直接写入，存在，从指定位置写入
    if sheet_name in workbook.sheetnames:
        start_row = workbook[sheet_name].max_row
    else:
        dateframe.to_excel(writer, sheet_name=sheet_name, index=False, header=True)
        writer.save()
        return
    sheet = workbook[sheet_name]  # 打开要编辑的工作表

    if start_row <= 1:
        #  sheet 已存在，直接追加会新增一个同名的sheet页，所以先复制再追加。
        writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)  # 复制已存在的sheet
        dateframe.to_excel(writer, sheet_name=sheet_name, index=False, header=True)
        writer.save()
        return

    for i in range(0, rows):
        for j in range(0, cols):  # value.shape[1]获得列数
            sheet.cell(row=start_row + 1 + i, column=j + 1, value=dateframe.iloc[i][j])
    workbook.save(path)


def hidden_sheet(path='D:\\test\\test.xlsx', sheet_name='服务信息'):
    """
    隐藏sheet
    :param path:
    :param sheet_name:
    :return:
    """
    workbook = load_workbook(path)  # 打开要写入数据的工作簿
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]  # 打开要编辑的工作表
        sheet.sheet_state = 'hidden'
    workbook.save(path)


def read_excel_by_row(path='D:\\test\\test.xlsx', sheet_name='Sheet1'):
    workbook = load_workbook(path)
    worksheet = workbook.get_sheet_by_name(sheet_name)
    for one in list(worksheet.rows):
        row = [item.value for item in one]
        print('行值', row)

    #
    # col3 = [item.value for item in list(worksheet.columns)[2]]
    #
    # print('第3行值', col3)
    #
    # cell_2_3 = worksheet.cell(row=2, column=3).value
    #
    # print('第2行第3列值', cell_2_3)
    #
    # max_row = worksheet.max_row
    #
    # print('最大行', max_row)


def get_map(path='D:\\test\\test.xlsx', sheet_name='Sheet1', k_col_index=3, v_col_index=4):
    """
    两列转map
    :param k_col_index:
    :param v_col_index:
    :return:
    """
    map = {}
    workbook = load_workbook(path)
    worksheet = workbook[sheet_name]
    for i in range(1, worksheet.max_row):
        key = worksheet.cell(row=i+1, column=k_col_index).value
        # val = worksheet.cell(row=i, column=v_col_index).value
        val = [item.value for item in list(worksheet.rows)[i]]
        if key is None:
            # map[key] = val 不会报错
            print('\033[4;33m' + '第【{}】行【{}】列为空，已跳过……'.format(i, k_col_index) + '\033[0m')
        else:
            if val is None:
                print('\033[4;33m' + '第【{}】行【{}】列为空'.format(i, v_col_index) + '\033[0m')
            map[key] = val

    print(map)
    return map


if __name__ == '__main__':
    # read_excel_by_row()
    map = get_map()
    for f in map:
        print(f, map[f][4])

